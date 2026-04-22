#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import re
import sys
import time
import json

from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import pyarrow.parquet as pq
import pythoncom
import polars as pl

from openpyxl import Workbook

os.environ.setdefault("POLARS_MAX_THREADS", "1")


PARQUET_NOME = "VIVO_TXTS_CONSOLIDADO.parquet"
TMP_NOME = "_tmp_shards_vivo"
CACHE_DIR = Path.home() / "AppData" / "Local" / "ValidadorVIVO"
CACHE_PARQUET = CACHE_DIR / "base_processada.parquet"
CACHE_META = CACHE_DIR / "base_processada_meta.json"
CACHE_ANDERSEN_PARQUET = CACHE_DIR / "base_andersen_conferencia.parquet"
CACHE_VIVO_PARQUET = CACHE_DIR / "base_vivo_conferencia.parquet"
CACHE_CONFERENCIA_META = CACHE_DIR / "conferencia_bases_meta.json"
CACHE_EXECUCOES_DIR = CACHE_DIR / "execucoes_conferencia"

LOGO_VIVO_ARQ = "logo_vivo.png"

COMPRESSION = "zstd"
ROW_GROUP = 100_000
PIPE_PLACEHOLDER = "<<<PIPE_DESC>>>"

ARQ_TABELA_DIV = "Tabela_Divisao.csv"
ARQ_CFOP_MAP = "Mapeamento_CFOP.csv"


# =========================
# RECURSOS
# =========================
def caminho_recurso(nome_arquivo):
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, nome_arquivo)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), nome_arquivo)


# =========================
# FUNÇÕES DO SCRIPT
# =========================

def descobrir_periodos_parquet(parquet_path):
    try:
        dfp = (
            pl.scan_parquet(str(parquet_path))
            .select(pl.col("Período").cast(pl.Utf8).fill_null("").alias("Período"))
            .unique()
            .collect()
        )
        vals = [x for x in dfp["Período"].to_list() if str(x).strip()]
        vals = sorted(set(vals))
        return vals
    except Exception:
        return []

def resumir_diretorio(path_str, max_len=40):
    s = str(path_str or "")
    if len(s) <= max_len:
        return s
    return s[:max_len-3] + "..."

def csv_para_xlsb_via_excel(csv_path, xlsb_path, sheet_name="Plan1", progress_callback=None, progresso_base=0, progresso_total=100):
    import win32com.client as win32

    pythoncom.CoInitialize()

    excel = None
    wb = None

    def report(passo, msg):
        if progress_callback:
            progress_callback("exportando_xlsx", passo, progresso_total, msg)

    try:
        csv_path = str(Path(csv_path).resolve())
        xlsb_path = str(Path(xlsb_path).resolve())

        report(progresso_base + 1, "Abrindo Excel...")
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        excel.AskToUpdateLinks = False
        excel.DisplayStatusBar = False
        excel.Interactive = False
        excel.UserControl = False

        try:
            excel.AutoRecover.Enabled = False
        except Exception:
            pass

        report(progresso_base + 2, "Abrindo CSV no Excel...")
        excel.Workbooks.OpenText(
            Filename=csv_path,
            Origin=65001,
            StartRow=1,
            DataType=1,
            TextQualifier=1,
            ConsecutiveDelimiter=False,
            Tab=False,
            Semicolon=True,
            Comma=False,
            Space=False,
            Other=False,
            Local=True
        )

        wb = excel.ActiveWorkbook

        report(progresso_base + 3, "Carregando dados na planilha...")
        try:
            wb.Worksheets(1).Name = sheet_name[:31]
        except Exception:
            pass

        report(progresso_base + 4, "Salvando arquivo XLSB...")
        wb.SaveAs(xlsb_path, FileFormat=50)

        report(progresso_base + 5, "Fechando arquivo...")
        wb.Close(False)
        wb = None

        excel.Quit()
        excel = None

        report(progresso_base + 6, "Conversão concluída.")

    finally:
        try:
            if wb is not None:
                wb.Close(False)
        except Exception:
            pass

        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass

        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def excel_col_name(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def escrever_bloco_excel(ws, row_ini, col_ini, dados):
    if not dados:
        return

    nrows = len(dados)
    ncols = len(dados[0]) if nrows else 0
    if ncols == 0:
        return

    col_fim = col_ini + ncols - 1
    row_fim = row_ini + nrows - 1

    a1 = f"{excel_col_name(col_ini)}{row_ini}"
    b2 = f"{excel_col_name(col_fim)}{row_fim}"

    ws.Range(a1, b2).Value = dados


def escrever_xlsx_via_excel_com(parquet_path, caminho_saida, sheet_name, processar_batch, progress_callback=None):
    import win32com.client as win32

    pythoncom.CoInitialize()

    excel = None
    wb = None

    try:
        parquet_path = str(parquet_path)
        caminho_saida = str(Path(caminho_saida).resolve())

        batch_size = 25_000
        max_excel_rows = 1_048_576

        pf = pq.ParquetFile(parquet_path)

        total_batches = 0
        if pf.metadata:
            for i in range(pf.metadata.num_row_groups):
                rg = pf.metadata.row_group(i)
                n = rg.num_rows
                total_batches += max(1, (n + batch_size - 1) // batch_size)

        if total_batches <= 0:
            total_batches = 1

        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        excel.AskToUpdateLinks = False

        wb = excel.Workbooks.Add()
        ws = wb.Worksheets(1)
        ws.Name = sheet_name[:31]

        header_escrito = False
        row_excel = 1
        linhas_lidas = 0
        linhas_gravadas = 0
        lote = 0

        for batch in pf.iter_batches(batch_size=batch_size, use_threads=True):
            linhas_lidas += batch.num_rows

            df = pl.from_arrow(batch)
            df = processar_batch(df)

            lote += 1

            if df.height == 0:
                if progress_callback:
                    progress_callback(
                        "exportando_xlsx",
                        lote,
                        total_batches,
                        f"Lote {lote}/{total_batches} | lidas: {linhas_lidas:,} | gravadas: {linhas_gravadas:,}"
                    )
                continue

            if not header_escrito:
                header = [str(c) for c in df.columns]
                escrever_bloco_excel(ws, 1, 1, [header])
                header_escrito = True
                row_excel = 2

            if row_excel + df.height - 1 > max_excel_rows:
                raise ValueError(
                    f"O Excel suporta no máximo {max_excel_rows:,} linhas por aba "
                    f"(incluindo cabeçalho). A exportação excedeu esse limite."
                )

            dados = df.rows()
            escrever_bloco_excel(ws, row_excel, 1, dados)

            row_excel += df.height
            linhas_gravadas += df.height

            if progress_callback:
                progress_callback(
                    "exportando_xlsx",
                    lote,
                    total_batches,
                    f"Lote {lote}/{total_batches} | lidas: {linhas_lidas:,} | gravadas: {linhas_gravadas:,}"
                )

        if not header_escrito:
            ws.Cells(1, 1).Value = ""

        wb.SaveAs(caminho_saida, FileFormat=51)
        wb.Close(False)
        wb = None

        excel.Quit()
        excel = None

        if progress_callback:
            progress_callback(
                "finalizado_xlsx",
                total_batches + 8,
                total_batches + 8,
                f"Exportação concluída | gravadas: {linhas_gravadas:,}"
            )

    finally:
        try:
            if wb is not None:
                wb.Close(False)
        except Exception:
            pass

        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass

        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def lista_unica(seq):
    vistos = set()
    out = []
    for x in seq:
        if x not in vistos:
            vistos.add(x)
            out.append(x)
    return out


def escrever_xlsx_bruto_texto(df, caminho_saida, nome_aba="BASE_VIVO"):
    caminho_saida = str(caminho_saida)

    max_excel_rows = 1_048_576
    total_rows_planilha = df.height + 1

    if total_rows_planilha > max_excel_rows:
        raise ValueError(
            f"O Excel suporta no máximo {max_excel_rows:,} linhas por aba. "
            f"A exportação teria {total_rows_planilha:,} linhas incluindo o cabeçalho."
        )

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=nome_aba[:31])

    ws.append([str(c) for c in df.columns])

    for fatia in df.iter_slices(n_rows=50_000):
        for row in fatia.iter_rows(named=False, buffer_size=50_000):
            ws.append(list(row))

    wb.save(caminho_saida)
    wb.close()


def limpar_colunas_csv(df):
    novas = []
    for c in df.columns:
        novo = (
            c.replace("ï»¿", "")
             .replace("Ã³", "ó")
             .replace("Ã£", "ã")
             .replace("Ã§", "ç")
             .replace("Ã©", "é")
             .replace("Ãª", "ê")
             .replace("Ã¡", "á")
             .replace("Ã­", "í")
             .replace("Ãº", "ú")
             .strip()
        )
        novas.append(novo)
    df.columns = novas
    return df


def carregar_divisoes_df():
    df = pl.read_csv(
        caminho_recurso(ARQ_TABELA_DIV),
        separator=";",
        encoding="utf8",
        ignore_errors=True
    )
    df = limpar_colunas_csv(df)

    return (
        df.select([
            pl.col("Local de Negócios").cast(pl.Utf8),
            pl.col("Divisão").cast(pl.Utf8).alias("Divisão_DePara"),
        ])
        .unique(subset=["Local de Negócios"], keep="first")
    )


def carregar_cfop_df():
    df = pl.read_csv(
        caminho_recurso(ARQ_CFOP_MAP),
        separator=";",
        encoding="utf8",
        ignore_errors=True
    )
    df = limpar_colunas_csv(df)

    return (
        df.select([
            pl.col("CFOP").cast(pl.Utf8).str.strip_chars().alias("CFOP"),
            pl.col("Mapeamento").cast(pl.Utf8).str.strip_chars().alias("Mapeamento"),
        ])
        .unique(subset=["CFOP"], keep="first")
    )


def detectar_header(path):
    with open(path, "r", encoding="latin-1", errors="ignore") as f:
        for i, linha in enumerate(f):
            up = linha.upper()
            if "|" in linha and ("CHAVE DA NOTA" in up or "MNFSM_CHV_NFE" in up):
                header = [c.strip() for c in linha.rstrip("\r\n").split("|")]
                return i, header
    return None, None


def limpar_nomes_colunas(cols):
    vistos = {}
    saida = []

    for c in cols:
        nome = c.strip() or "COLUNA_VAZIA"
        if nome in vistos:
            vistos[nome] += 1
            nome = f"{nome}_{vistos[nome]}"
        else:
            vistos[nome] = 1
        saida.append(nome)

    return saida


_REGEX_LINHAS_SELECIONADAS = re.compile(
    r"\d+\s+linhas\s+selecionadas\.?",
    flags=re.IGNORECASE,
)


def linha_eh_lixo(linha):
    s = linha.strip()
    if not s:
        return True

    # `s.strip("-|")` é muito mais rápido que `set(s) <= {"-","|"}`: evita
    # construir um set e retorna string vazia só pra linhas de separador.
    if not s.strip("-|"):
        return True

    # Header repetido no meio do arquivo começa com "CHAVE"/"MNFSM" (alfa).
    # Linhas de dados começam com dígito (chave NFe de 44 caracteres). Um
    # check de 1 char no primeiro caractere descarta 99% sem precisar chamar
    # `s.upper()` na linha inteira.
    first = s[0]
    if first.isalpha():
        up = s.upper()
        if "MNFSM_CHV_NFE" in up or "CHAVE DA NOTA" in up:
            return True

    # Rodapé "N linhas selecionadas.": pré-filtro barato por substring antes
    # do regex (evita custo em todas as linhas de dados normais).
    if "linhas" in s and _REGEX_LINHAS_SELECIONADAS.fullmatch(s):
        return True

    return False


def descobrir_idx_dsc(header_cols):
    prioridades = [
        "DSC",
        "INFSM_DSC",
        "DSC_1",
        "DESCRICAO",
        "DESCRIÇÃO",
    ]

    upper_cols = [c.upper() for c in header_cols]

    for alvo in prioridades:
        if alvo in upper_cols:
            return upper_cols.index(alvo)

    for i, c in enumerate(upper_cols):
        if "DSC" in c:
            return i

    return None


def descobrir_idx_merge_seguro(header_cols):
    """Encontra uma coluna de texto livre *tardia* no header, ideal para
    absorver os pipes extras sem corromper colunas numéricas críticas
    (VAL_ICMS, BAS_ICMS, ALIQ_ICMS, etc).

    O SAP injeta pipes literais dentro de campos de audit log como VAR05,
    VAR04, OBSERVACAO — esses campos já são free-text e vêm depois dos
    valores fiscais, então mesclar extras neles preserva o alinhamento
    de VAL_ICMS e evita que o DSC-merging (hardcoded pra coluna 24) empurre
    CFOP/NCM/NOPE pra dentro do DSC e shifte todas as colunas subsequentes.

    Retorna None se nenhuma coluna segura for encontrada (cai no fallback
    de mesclar no DSC, comportamento antigo).
    """
    prioridades = ["VAR05", "VAR04", "VAR03", "VAR02", "VAR01"]
    upper_cols = [c.upper() for c in header_cols]
    for alvo in prioridades:
        if alvo in upper_cols:
            return upper_cols.index(alvo)
    return None


def corrigir_pipe_na_descricao(linha, ncols, idx_dsc):
    linha = linha.rstrip("\r\n")

    # Fast path: conta `|` sem fazer split. Se bater com o esperado, a linha
    # não precisa ser modificada — devolve a string original direto.
    # Isso pula o split+join custoso em ~99% das linhas (as correctamente
    # formatadas), que são a grande maioria em volume.
    if linha.count("|") == ncols - 1:
        return linha

    cols = linha.split("|")

    if idx_dsc is None:
        if len(cols) < ncols:
            cols.extend([""] * (ncols - len(cols)))
        elif len(cols) > ncols:
            cols = cols[:ncols]
        return "|".join(cols)

    if len(cols) < ncols:
        cols.extend([""] * (ncols - len(cols)))
        return "|".join(cols)

    right_count = ncols - idx_dsc - 1

    left = cols[:idx_dsc]

    if right_count > 0:
        right = cols[-right_count:]
        middle = cols[idx_dsc:-right_count]
    else:
        right = []
        middle = cols[idx_dsc:]

    dsc = PIPE_PLACEHOLDER.join(middle)

    cols_corrigidas = left + [dsc] + right

    if len(cols_corrigidas) < ncols:
        cols_corrigidas.extend([""] * (ncols - len(cols_corrigidas)))
    elif len(cols_corrigidas) > ncols:
        cols_corrigidas = cols_corrigidas[:ncols]

    return "|".join(cols_corrigidas)


def extrair_divisao_arquivo(nome):
    nome_up = (nome or "").upper()

    if "0001SP" in nome_up:
        return "29SP"

    m = re.search(r"NFE_([0-9]{2}[A-Z]{2})", nome_up)
    if m:
        return m.group(1)

    return ""


def montar_ordem_final(cols):
    remover = {"DivArquivo", "Divisão_DePara", "CFOP", "__ordem__"}

    base = [c for c in cols if c not in remover]

    for c in ["Fonte", "Período", "Nome do Arquivo", "Divisão", "Mapeamento"]:
        if c in base:
            base.remove(c)

    ordem = ["Fonte", "Período", "Nome do Arquivo"]

    for c in base:
        ordem.append(c)

        if c == "EMPRESA" and "Divisão" in cols and "Divisão" not in ordem:
            ordem.append("Divisão")

        if c == "CFOP_COD" and "Mapeamento" in cols and "Mapeamento" not in ordem:
            ordem.append("Mapeamento")

    for c in ["Divisão", "Mapeamento"]:
        if c in cols and c not in ordem:
            ordem.append(c)

    return ordem


def detectar_tipo_movimento(arquivos):
    tem_ent = False
    tem_sai = False

    for arq in arquivos:
        nome = Path(arq).name.upper()

        if "ENT" in nome:
            tem_ent = True

        if "SAI" in nome or "SAIDA" in nome:
            tem_sai = True

    if tem_ent and tem_sai:
        raise ValueError(
            "A pasta contém arquivos de ENTRADA e SAÍDA juntos. "
            "Processe apenas um tipo por vez."
        )

    if tem_ent:
        return "ENTRADA"

    if tem_sai:
        return "SAIDA"

    raise ValueError(
        "Não foi possível identificar se os arquivos são de ENTRADA ou SAÍDA "
        "pelos nomes dos TXT."
    )


def salvar_meta_cache(tipo_movimento, base_dir, pasta_destino):
    CACHE_DIR.mkdir(parents=True, exist_ok=True)

    meta = {
        "tipo_movimento": tipo_movimento,
        "base_dir": str(base_dir),
        "pasta_destino": str(pasta_destino),
        "cache_parquet": str(CACHE_PARQUET),
    }

    with open(CACHE_META, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)


def carregar_meta_cache():
    if not CACHE_META.exists():
        return None

    with open(CACHE_META, "r", encoding="utf-8") as f:
        return json.load(f)

def expr_numero_br(col_name):
    s = pl.col(col_name).cast(pl.Utf8).fill_null("").str.strip_chars()

    # trata negativo no final: 123,45-
    s = (
        pl.when(s.str.ends_with("-"))
        .then(pl.lit("-") + s.str.slice(0, s.str.len_chars() - 1))
        .otherwise(s)
    )

    # remove separador de milhar e troca vírgula por ponto
    s = (
        s.str.replace_all(r"\.", "")
         .str.replace(",", ".", literal=True)
    )

    return s.cast(pl.Float64, strict=False).fill_null(0.0)

def montar_base_vivo_conferencia_filtrada(lf, tipo_movimento):
    cols = lf.collect_schema().names()
    tipo = (tipo_movimento or "").upper()

    def norm_upper(c):
        return (
            pl.col(c)
            .cast(pl.Utf8)
            .fill_null("")
            .str.strip_chars()
            .str.to_uppercase()
        )

    def norm_txt(c):
        return (
            pl.col(c)
            .cast(pl.Utf8)
            .fill_null("")
            .str.strip_chars()
        )

    exprs_base = []

    if "Divisão" in cols:
        exprs_base.append(
            pl.col("Divisão").cast(pl.Utf8).fill_null("").str.strip_chars().str.to_uppercase().alias("Divisão")
        )
    else:
        exprs_base.append(pl.lit("").alias("Divisão"))

    if "CFOP_COD" in cols:
        exprs_base.append(
            pl.col("CFOP_COD").cast(pl.Utf8).fill_null("").str.strip_chars().alias("CFOP")
        )
    else:
        exprs_base.append(pl.lit("").alias("CFOP"))

    if "Mapeamento" in cols:
        exprs_base.append(
            pl.col("Mapeamento").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Mapeamento")
        )
    else:
        exprs_base.append(pl.lit("").alias("Mapeamento"))

    if "Período" in cols:
        exprs_base.append(pl.col("Período").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Período"))
    else:
        exprs_base.append(pl.lit("").alias("Período"))

    if "Nome do Arquivo" in cols:
        exprs_base.append(pl.col("Nome do Arquivo").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Nome do Arquivo"))
    else:
        exprs_base.append(pl.lit("").alias("Nome do Arquivo"))

    if tipo == "ENTRADA":
        excluir_cfop = ["1923", "2923", "1915", "2915", "1154", "2154", "1403", "2403", "1555", "2555"]
        col_trib = "TRIBICMS" if "TRIBICMS" in cols else ("TRIBICM" if "TRIBICM" in cols else None)

        exprs_norm = []
        if "CFOP_COD" in cols:
            exprs_norm.append(norm_upper("CFOP_COD").alias("CFOP_COD"))
        if "IND_CANC" in cols:
            exprs_norm.append(norm_upper("IND_CANC").alias("IND_CANC"))
        if col_trib:
            exprs_norm.append(norm_upper(col_trib).alias(col_trib))
        if "Mapeamento" in cols:
            exprs_norm.append(norm_txt("Mapeamento").alias("Mapeamento"))
        if "Divisão" in cols:
            exprs_norm.append(
                pl.col("Divisão").cast(pl.Utf8).fill_null("").str.strip_chars().str.to_uppercase().alias("Divisão")
            )

        lf2 = lf.with_columns(exprs_norm)

        filtros = []
        if col_trib:
            filtros.append(pl.col(col_trib) == "S")
        if "IND_CANC" in cols:
            filtros.append(pl.col("IND_CANC") == "N")
        if "CFOP_COD" in cols:
            filtros.append(~pl.col("CFOP_COD").is_in(excluir_cfop))

        if filtros:
            expr = filtros[0]
            for f in filtros[1:]:
                expr = expr & f
            lf2 = lf2.filter(expr)

        valor_expr = expr_numero_br("VAL_ICMS") if "VAL_ICMS" in cols else pl.lit(0.0)

        return (
            lf2.with_columns(exprs_base + [
                pl.lit("Vivo").alias("Base"),
                pl.lit("Livro de Entrada").alias("Fonte"),
                pl.lit("Entrada").alias("Tipo"),
                valor_expr.alias("Valor_ICMS_Conf"),
            ])
            .select([
                "Base",
                "Fonte",
                "Período",
                "Nome do Arquivo",
                "Divisão",
                "CFOP",
                "Mapeamento",
                "Tipo",
                "Valor_ICMS_Conf",
            ])
        )

    elif tipo == "SAIDA":
        exprs_norm = []
        if "CFOP_COD" in cols:
            exprs_norm.append(norm_upper("CFOP_COD").alias("CFOP_COD"))
        if "IND_CANC" in cols:
            exprs_norm.append(norm_upper("IND_CANC").alias("IND_CANC"))
        if "I_2" in cols:
            exprs_norm.append(norm_upper("I_2").alias("I_2"))
        if "Mapeamento" in cols:
            exprs_norm.append(norm_txt("Mapeamento").alias("Mapeamento"))
        if "Divisão" in cols:
            exprs_norm.append(
                pl.col("Divisão").cast(pl.Utf8).fill_null("").str.strip_chars().str.to_uppercase().alias("Divisão")
            )

        lf2 = lf.with_columns(exprs_norm)

        filtros = []
        if "I_2" in cols:
            filtros.append(pl.col("I_2") == "S")
        if "IND_CANC" in cols:
            filtros.append(pl.col("IND_CANC") == "N")

        if filtros:
            expr = filtros[0]
            for f in filtros[1:]:
                expr = expr & f
            lf2 = lf2.filter(expr)

        valor_expr = (
            expr_numero_br("INFSM_VAL_ICMS")
            if "INFSM_VAL_ICMS" in cols else
            expr_numero_br("VAL_ICMS")
            if "VAL_ICMS" in cols else
            pl.lit(0.0)
        )

        return (
            lf2.with_columns(exprs_base + [
                pl.lit("Vivo").alias("Base"),
                pl.lit("Livro de Saída").alias("Fonte"),
                pl.lit("Saída").alias("Tipo"),
                valor_expr.alias("Valor_ICMS_Conf"),
            ])
            .select([
                "Base",
                "Fonte",
                "Período",
                "Nome do Arquivo",
                "Divisão",
                "CFOP",
                "Mapeamento",
                "Tipo",
                "Valor_ICMS_Conf",
            ])
        )

    else:
        raise ValueError(f"Tipo de movimento inválido para conferência: {tipo_movimento}")


def gerar_bases_conferencia_cache(parquet_path, tipo_movimento, base_dir):
    CACHE_EXECUCOES_DIR.mkdir(parents=True, exist_ok=True)

    parquet_path = str(parquet_path)
    tipo = (tipo_movimento or "").upper()

    periodos = descobrir_periodos_parquet(parquet_path)
    if len(periodos) == 1:
        periodo_txt = periodos[0]
    elif len(periodos) > 1:
        periodo_txt = f"MULTI_{len(periodos)}_PERIODOS"
    else:
        periodo_txt = "SEM_PERIODO"

    exec_id = time.strftime("%Y%m%d_%H%M%S")
    exec_dir = CACHE_EXECUCOES_DIR / f"{exec_id}_{tipo}"
    exec_dir.mkdir(parents=True, exist_ok=True)

    nome_copia = f"BASE_INTERNA__{periodo_txt}__{tipo}.parquet"
    copia_nomeada = exec_dir / nome_copia

    (
        pl.scan_parquet(parquet_path)
        .sink_parquet(str(copia_nomeada), compression=COMPRESSION)
    )

    cache_andersen = exec_dir / "base_andersen_conferencia.parquet"
    cache_vivo = exec_dir / "base_vivo_conferencia.parquet"
    cache_meta = exec_dir / "meta.json"

    lf = pl.scan_parquet(parquet_path)
    cols = lf.collect_schema().names()

    # =========================
    # BASE ANDERSEN CONFERÊNCIA
    # =========================
    cols_andersen = [c for c in [
        "Fonte", "Período", "Nome do Arquivo", "Divisão", "CFOP_COD", "Mapeamento",
        "VAL_ICMS", "INFSM_VAL_ICMS", "IND_CANC", "TRIBICM", "TRIBICMS", "I_2"
    ] if c in cols]

    if cols_andersen:
        lf_and = lf.select(cols_andersen)

        if tipo == "ENTRADA":
            col_trib = "TRIBICMS" if "TRIBICMS" in cols else ("TRIBICM" if "TRIBICM" in cols else None)
            excluir_cfop = ["1923", "2923", "1915", "2915", "1154", "2154", "1403", "2403", "1555", "2555"]

            exprs_norm = []
            if "CFOP_COD" in cols:
                exprs_norm.append(
                    pl.col("CFOP_COD").cast(pl.Utf8).fill_null("").str.strip_chars().str.to_uppercase().alias("CFOP_COD")
                )
            if "IND_CANC" in cols:
                exprs_norm.append(
                    pl.col("IND_CANC").cast(pl.Utf8).fill_null("").str.strip_chars().str.to_uppercase().alias("IND_CANC")
                )
            if col_trib:
                exprs_norm.append(
                    pl.col(col_trib).cast(pl.Utf8).fill_null("").str.strip_chars().str.to_uppercase().alias(col_trib)
                )

            lf_and = lf_and.with_columns(exprs_norm)

            filtros = []
            if col_trib:
                filtros.append(pl.col(col_trib) == "S")
            if "IND_CANC" in cols:
                filtros.append(pl.col("IND_CANC") == "N")
            if "CFOP_COD" in cols:
                filtros.append(~pl.col("CFOP_COD").is_in(excluir_cfop))

            if filtros:
                expr = filtros[0]
                for f in filtros[1:]:
                    expr = expr & f
                lf_and = lf_and.filter(expr)

            valor_expr_and = expr_numero_br("VAL_ICMS") if "VAL_ICMS" in cols else pl.lit(0.0)

            (
                lf_and.with_columns([
                    pl.lit("Andersen").alias("Base"),
                    pl.lit("Andersen").alias("Fonte"),
                    pl.col("Período").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Período")
                        if "Período" in cols else pl.lit("").alias("Período"),
                    pl.col("Nome do Arquivo").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Nome do Arquivo")
                        if "Nome do Arquivo" in cols else pl.lit("").alias("Nome do Arquivo"),
                    pl.col("Divisão").cast(pl.Utf8).fill_null("").str.strip_chars().str.to_uppercase().alias("Divisão")
                        if "Divisão" in cols else pl.lit("").alias("Divisão"),
                    pl.col("CFOP_COD").cast(pl.Utf8).fill_null("").str.strip_chars().alias("CFOP")
                        if "CFOP_COD" in cols else pl.lit("").alias("CFOP"),
                    pl.col("Mapeamento").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Mapeamento")
                        if "Mapeamento" in cols else pl.lit("").alias("Mapeamento"),
                    pl.lit("Entrada").alias("Tipo"),
                    valor_expr_and.alias("Valor_ICMS_Conf"),
                ])
                .select([
                    "Base",
                    "Fonte",
                    "Período",
                    "Nome do Arquivo",
                    "Divisão",
                    "CFOP",
                    "Mapeamento",
                    "Tipo",
                    "Valor_ICMS_Conf",
                ])
                .sink_parquet(str(cache_andersen), compression=COMPRESSION)
            )

        elif tipo == "SAIDA":
            exprs_norm = []
            if "CFOP_COD" in cols:
                exprs_norm.append(
                    pl.col("CFOP_COD").cast(pl.Utf8).fill_null("").str.strip_chars().str.to_uppercase().alias("CFOP_COD")
                )
            if "IND_CANC" in cols:
                exprs_norm.append(
                    pl.col("IND_CANC").cast(pl.Utf8).fill_null("").str.strip_chars().str.to_uppercase().alias("IND_CANC")
                )
            if "I_2" in cols:
                exprs_norm.append(
                    pl.col("I_2").cast(pl.Utf8).fill_null("").str.strip_chars().str.to_uppercase().alias("I_2")
                )

            lf_and = lf_and.with_columns(exprs_norm)

            filtros = []
            if "I_2" in cols:
                filtros.append(pl.col("I_2") == "S")
            if "IND_CANC" in cols:
                filtros.append(pl.col("IND_CANC") == "N")

            if filtros:
                expr = filtros[0]
                for f in filtros[1:]:
                    expr = expr & f
                lf_and = lf_and.filter(expr)

            valor_expr_and = (
                expr_numero_br("INFSM_VAL_ICMS")
                if "INFSM_VAL_ICMS" in cols else
                expr_numero_br("VAL_ICMS")
                if "VAL_ICMS" in cols else
                pl.lit(0.0)
            )

            (
                lf_and.with_columns([
                    pl.lit("Andersen").alias("Base"),
                    pl.lit("Andersen").alias("Fonte"),
                    pl.col("Período").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Período")
                        if "Período" in cols else pl.lit("").alias("Período"),
                    pl.col("Nome do Arquivo").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Nome do Arquivo")
                        if "Nome do Arquivo" in cols else pl.lit("").alias("Nome do Arquivo"),
                    pl.col("Divisão").cast(pl.Utf8).fill_null("").str.strip_chars().str.to_uppercase().alias("Divisão")
                        if "Divisão" in cols else pl.lit("").alias("Divisão"),
                    pl.col("CFOP_COD").cast(pl.Utf8).fill_null("").str.strip_chars().alias("CFOP")
                        if "CFOP_COD" in cols else pl.lit("").alias("CFOP"),
                    pl.col("Mapeamento").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Mapeamento")
                        if "Mapeamento" in cols else pl.lit("").alias("Mapeamento"),
                    pl.lit("Saída").alias("Tipo"),
                    valor_expr_and.alias("Valor_ICMS_Conf"),
                ])
                .select([
                    "Base",
                    "Fonte",
                    "Período",
                    "Nome do Arquivo",
                    "Divisão",
                    "CFOP",
                    "Mapeamento",
                    "Tipo",
                    "Valor_ICMS_Conf",
                ])
                .sink_parquet(str(cache_andersen), compression=COMPRESSION)
            )

    # =========================
    # BASE VIVO CONFERÊNCIA
    # =========================
    base_vivo_conf = montar_base_vivo_conferencia_filtrada(lf, tipo)
    base_vivo_conf.sink_parquet(str(cache_vivo), compression=COMPRESSION)

    meta = {
        "exec_id": exec_id,
        "periodo": periodo_txt,
        "tipo_movimento": tipo,
        "base_dir": str(base_dir),
        "base_dir_resumido": resumir_diretorio(base_dir),
        "data_execucao": time.strftime("%Y-%m-%d %H:%M:%S"),
        "andersen": str(cache_andersen),
        "vivo": str(cache_vivo),
        "base_processada": str(copia_nomeada),
        "label": f"{periodo_txt} - {'Entrada' if tipo == 'ENTRADA' else 'Saída'} | {resumir_diretorio(base_dir)}"
    }

    with open(cache_meta, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

    with open(CACHE_CONFERENCIA_META, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

def selecionar_colunas_existentes(df, colunas_desejadas):
    cols_existentes = df.columns
    ordem_final = [c for c in colunas_desejadas if c in cols_existentes]
    return df.select(ordem_final)


def renomear_colunas_parcial(df, mapa_renomeio):
    ren = {k: v for k, v in mapa_renomeio.items() if k in df.columns}
    if ren:
        df = df.rename(ren)
    return df


def _extrair_periodo_do_parquet(parquet_path):
    """Extrai o período do nome do parquet (BASE_INTERNA__{periodo}__{tipo}.parquet)."""
    nome = Path(parquet_path).stem
    partes = nome.split("__")
    if len(partes) >= 2:
        return partes[1]
    return ""


def _build_text_column_formats(columns):
    """Retorna dict de column_formats para colunas sensíveis (CNPJ, IE, Chave, etc.)"""
    TEXT_KEYWORDS = ["CNPJ", "CPF", "IE", "CHAVE", "SERIE", "CHV_NFE",
                     "Inscri", "NUM_FCI", "COD_CONT", "Chave"]
    fmt = {}
    for col_name in columns:
        for kw in TEXT_KEYWORDS:
            if kw.upper() in col_name.upper():
                fmt[col_name] = {"num_format": "@"}
                break
    return fmt


def _csv_para_xlsb_rapido(csv_path, xlsb_path, parquet_path=None,
                           text_col_info=None, progress_callback=None):
    """CSV -> XLSB via Excel COM.
    Apos abrir o CSV, corrige colunas texto (CNPJ/IE/Chave) relendo do parquet
    e colando como texto nativo no XLSB."""
    import win32com.client as win32

    pythoncom.CoInitialize()
    excel = None
    wb = None

    steps_total = 5
    step = [0]

    def report(msg):
        step[0] += 1
        if progress_callback:
            progress_callback("convertendo_xlsb", step[0], steps_total, msg)

    try:
        csv_path = str(Path(csv_path).resolve())
        xlsb_path = str(Path(xlsb_path).resolve())

        report("Abrindo Excel...")
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        excel.AskToUpdateLinks = False

        try:
            excel.AutoRecover.Enabled = False
        except Exception:
            pass

        report("Lendo CSV...")
        excel.Workbooks.OpenText(
            Filename=csv_path,
            Origin=65001,
            StartRow=1,
            DataType=1,
            TextQualifier=1,
            ConsecutiveDelimiter=False,
            Tab=False,
            Semicolon=True,
            Comma=False,
            Space=False,
            Other=False,
            Local=True,
        )

        wb = excel.ActiveWorkbook
        ws = wb.Worksheets(1)
        ws.Name = "Dados"

        # Corrigir colunas texto relendo do parquet
        if parquet_path and text_col_info:
            report("Corrigindo colunas texto (CNPJ/IE/Chave)...")
            last_row = ws.UsedRange.Rows.Count

            for col_name, excel_col_idx in text_col_info:
                try:
                    # Ler coluna do parquet
                    col_data = (
                        pl.scan_parquet(str(parquet_path))
                        .select(pl.col(col_name).cast(pl.Utf8).fill_null("").str.strip_chars())
                        .collect()
                    )[col_name].to_list()

                    # Formatar coluna como texto
                    col_letter = _col_letter(excel_col_idx)
                    rng = ws.Range(f"{col_letter}2:{col_letter}{last_row}")
                    rng.NumberFormat = "@"

                    # Colar valores como tupla de tuplas (1 coluna)
                    values = tuple((str(v),) for v in col_data)
                    rng.Value = values

                    del col_data, values
                except Exception as e:
                    report(f"  Aviso coluna {col_name}: {e}")

        # Monitoramento de progresso em tempo real durante SaveAs
        import threading

        csv_size = os.path.getsize(csv_path) if os.path.exists(csv_path) else 0
        estimated_xlsb = max(csv_size * 0.22, 10 * 1024 * 1024)

        monitor_stop = threading.Event()

        def _monitor():
            last_pct = 0
            while not monitor_stop.is_set():
                monitor_stop.wait(timeout=1.5)
                try:
                    if os.path.exists(xlsb_path):
                        sz = os.path.getsize(xlsb_path)
                        pct = min(int((sz / estimated_xlsb) * 100), 99)
                        if pct > last_pct:
                            last_pct = pct
                            mb = sz / (1024 * 1024)
                            if progress_callback:
                                progress_callback(
                                    "convertendo_xlsb",
                                    pct, 100,
                                    f"Salvando XLSB... {pct}% ({mb:.0f} MB)"
                                )
                except Exception:
                    pass

        monitor_thread = threading.Thread(target=_monitor, daemon=True)
        monitor_thread.start()

        report("Salvando XLSB...")
        wb.SaveAs(xlsb_path, FileFormat=50)

        monitor_stop.set()
        monitor_thread.join(timeout=2)

        final_mb = os.path.getsize(xlsb_path) / (1024*1024) if os.path.exists(xlsb_path) else 0
        if progress_callback:
            progress_callback("convertendo_xlsb", 100, 100,
                              f"XLSB salvo! ({final_mb:.0f} MB)")

        wb.Close(False)
        wb = None
        excel.Quit()
        excel = None

    finally:
        try:
            if wb is not None:
                wb.Close(False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _col_letter(n):
    """Converte indice 0-based para letra Excel (0=A, 25=Z, 26=AA, etc.)"""
    result = ""
    n += 1
    while n > 0:
        n -= 1
        result = chr(65 + n % 26) + result
        n //= 26
    return result


def _exportar_direto_excel(parquet_path, output_path, text_col_indices,
                           processar_batch=None, file_format=50,
                           batch_size=250_000, progress_callback=None):
    """Injeta dados direto na memoria do Excel via Range.Value (sem CSV intermediario).
    XLSB binario. Colunas texto pre-formatadas antes da injecao."""
    import win32com.client as win32

    pythoncom.CoInitialize()
    excel = None
    wb = None

    def report(msg):
        if progress_callback:
            progress_callback("exportando_xlsb", 1, 1, msg)

    try:
        output_path = str(Path(output_path).resolve())
        text_set = set(text_col_indices or [])

        report("Iniciando Excel...")
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        excel.AskToUpdateLinks = False
        try:
            excel.Calculation = -4135  # xlCalculationManual
        except Exception:
            pass

        try:
            excel.AutoRecover.Enabled = False
        except Exception:
            pass

        wb = excel.Workbooks.Add()
        ws = wb.Worksheets(1)
        ws.Name = "Dados"

        pf = pq.ParquetFile(parquet_path)
        row_cursor = 1  # 1-based (row 1 = header)
        header_written = False
        total_written = 0
        col_count = 0
        last_col_letter = ""

        for batch in pf.iter_batches(batch_size=batch_size, use_threads=True):
            df = pl.from_arrow(batch)

            if processar_batch:
                df = processar_batch(df)
                if df.height == 0:
                    continue

            df = df.with_columns([
                pl.col(c).str.strip_chars().alias(c)
                for c in df.columns
                if df.schema[c] == pl.Utf8
            ])

            columns = df.columns
            col_count = len(columns)
            last_col_letter = _col_letter(col_count - 1)

            if not header_written:
                # Pre-formatar colunas texto ANTES de injetar dados
                report("Formatando colunas texto...")
                for ci in text_set:
                    if ci < col_count:
                        try:
                            ws.Columns(ci + 1).NumberFormat = "@"
                        except Exception:
                            pass

                # Header
                header_tuple = tuple(columns)
                ws.Range(f"A1:{last_col_letter}1").Value = [header_tuple]
                row_cursor = 2
                header_written = True

            # Converter dados para tupla de tuplas (transferencia COM rapida)
            report(f"Injetando dados... ({total_written + df.height:,} linhas)")

            # Para colunas texto, converter para string antes
            pdf = df.to_pandas()
            for ci in text_set:
                if ci < col_count:
                    col_name = columns[ci]
                    if col_name in pdf.columns:
                        pdf[col_name] = pdf[col_name].fillna("").astype(str)

            # Substituir NaN por None (Excel blank) e converter
            pdf = pdf.where(pdf.notna(), None)
            data = tuple(tuple(row) for row in pdf.itertuples(index=False, name=None))

            end_row = row_cursor + len(data) - 1
            ws.Range(f"A{row_cursor}:{last_col_letter}{end_row}").Value = data
            row_cursor = end_row + 1
            total_written += len(data)

            del pdf, data

        if total_written == 0:
            wb.Close(False)
            wb = None
            excel.Quit()
            excel = None
            # Criar arquivo vazio
            pl.DataFrame().write_excel(workbook=output_path, worksheet="Dados")
            return 0

        fmt_name = "XLSB" if file_format == 50 else "XLSX"
        report(f"Salvando {fmt_name} ({total_written:,} linhas)...")
        wb.SaveAs(output_path, FileFormat=file_format)

        wb.Close(False)
        wb = None
        excel.Quit()
        excel = None

        report(f"Concluido! {total_written:,} linhas")
        return total_written

    finally:
        try:
            if wb is not None:
                wb.Close(False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _detect_text_column_indices(columns):
    """Retorna lista de indices (0-based) das colunas que devem ser texto."""
    TEXT_KEYWORDS = ["CNPJ", "CPF", "IE", "CHAVE", "SERIE", "CHV_NFE",
                     "Inscri", "NUM_FCI", "COD_CONT", "Chave"]
    indices = []
    for i, col_name in enumerate(columns):
        for kw in TEXT_KEYWORDS:
            if kw.upper() in col_name.upper():
                indices.append(i)
                break
    return indices


def exportar_versao_andersen(parquet_path, pasta_destino, tipo_movimento=None, progress_callback=None):
    """Pipeline nuclear: CSV (Polars/Rust, ~10s) -> XLSX (Excel COM/C++, ~30s).
    Total esperado: <1 minuto para 900k+ linhas x 116 colunas."""
    pasta_destino = Path(pasta_destino)
    tipo_label = (tipo_movimento or "").strip().upper()
    if tipo_label == "ENTRADA":
        tipo_label = "Entrada"
    elif tipo_label == "SAIDA":
        tipo_label = "Saída"
    periodo = _extrair_periodo_do_parquet(parquet_path)
    sufixo = ""
    if tipo_label and periodo:
        sufixo = f" - {tipo_label}_{periodo}"
    elif tipo_label:
        sufixo = f" - {tipo_label}"
    elif periodo:
        sufixo = f" - {periodo}"

    out_xlsb = pasta_destino / f"BASE_ANDERSEN{sufixo}.xlsb"
    tmp_csv = pasta_destino / f"_tmp_andersen{sufixo}.csv"

    # PASSO 1: CSV com Polars (Rust nativo, ~10s)
    if progress_callback:
        progress_callback("exportando_csv", 1, 3, "Gerando CSV (Polars/Rust)...")

    pf = pq.ParquetFile(parquet_path)
    writer_iniciado = False
    linhas_gravadas = 0
    text_col_names_a = None

    with open(tmp_csv, "w", encoding="utf-8", newline="") as f:
        f.write("\ufeff")

        for batch in pf.iter_batches(batch_size=250_000, use_threads=True):
            df = pl.from_arrow(batch)

            df = df.with_columns([
                pl.col(c).str.strip_chars().alias(c)
                for c in df.columns
                if df.schema[c] == pl.Utf8
            ])

            if text_col_names_a is None:
                idx = _detect_text_column_indices(df.columns)
                text_col_names_a = [df.columns[i] for i in idx if i < len(df.columns)]

            df.write_csv(f, separator=";", include_header=not writer_iniciado,
                         null_value="", line_terminator="\n", quote_style="necessary")
            writer_iniciado = True
            linhas_gravadas += df.height

    if progress_callback:
        progress_callback("exportando_csv", 2, 3,
                          f"CSV pronto ({linhas_gravadas:,}). Convertendo XLSB...")

    # Montar info das colunas texto (nome_parquet, indice_excel)
    text_info = []
    if text_col_names_a:
        # Pegar header do CSV pra achar indices
        with open(tmp_csv, "r", encoding="utf-8-sig") as hf:
            csv_header = hf.readline().strip().split(";")
        for col_name in text_col_names_a:
            if col_name in csv_header:
                text_info.append((col_name, csv_header.index(col_name)))

    # PASSO 2: CSV -> XLSB + corrigir colunas texto via parquet
    try:
        _csv_para_xlsb_rapido(str(tmp_csv), str(out_xlsb),
                              parquet_path=parquet_path,
                              text_col_info=text_info,
                              progress_callback=progress_callback)
    finally:
        try:
            tmp_csv.unlink()
        except Exception:
            pass

    if progress_callback:
        progress_callback("finalizado_xlsb", 3, 3,
                          f"XLSB concluido | {linhas_gravadas:,} linhas")

    return out_xlsb


def exportar_versao_vivo(parquet_path, pasta_destino, tipo_movimento, progress_callback=None):
    tipo = (tipo_movimento or "").upper()
    parquet_path = str(parquet_path)
    pasta_destino = Path(pasta_destino)

    def normalizar_colunas_filtro(df, cols):
        exprs = []
        for c in cols:
            if c in df.columns:
                exprs.append(
                    pl.col(c)
                    .cast(pl.Utf8)
                    .fill_null("")
                    .str.strip_chars()
                    .str.to_uppercase()
                    .alias(c)
                )
        if exprs:
            df = df.with_columns(exprs)
        return df

    def escrever_xlsb_csv_com(parquet_path, xlsb_out, processar_batch):
        """CSV (Polars/Rust) + XLSB (Excel COM). Mais rapido que Range.Value."""
        import tempfile

        if progress_callback:
            progress_callback("exportando_csv", 1, 3, "Gerando CSV (Polars/Rust)...")

        tmp_csv = Path(tempfile.gettempdir()) / f"_tmp_vivo_{os.getpid()}.csv"
        pf_inner = pq.ParquetFile(parquet_path)
        writer_ok = False
        linhas = 0
        text_col_names = None
        text_chunks = []

        with open(tmp_csv, "w", encoding="utf-8", newline="") as f:
            f.write("\ufeff")

            for batch in pf_inner.iter_batches(batch_size=200_000, use_threads=True):
                df = pl.from_arrow(batch)
                df = processar_batch(df)
                if df.height == 0:
                    continue

                df = df.with_columns([
                    pl.col(c).str.strip_chars().alias(c)
                    for c in df.columns if df.schema[c] == pl.Utf8
                ])

                if text_col_names is None:
                    idx = _detect_text_column_indices(df.columns)
                    text_col_names = [df.columns[i] for i in idx if i < len(df.columns)]

                # Guardar valores texto pra corrigir depois
                if text_col_names:
                    existing = [c for c in text_col_names if c in df.columns]
                    if existing:
                        text_chunks.append(df.select([
                            pl.col(c).cast(pl.Utf8).fill_null("").str.strip_chars()
                            for c in existing
                        ]))

                df.write_csv(f, separator=";", include_header=not writer_ok,
                             null_value="", line_terminator="\n", quote_style="necessary")
                writer_ok = True
                linhas += df.height

        if not writer_ok:
            open(str(xlsb_out), "w").close()
            try:
                tmp_csv.unlink()
            except Exception:
                pass
            return 0, 1

        if progress_callback:
            progress_callback("exportando_csv", 2, 3,
                              f"CSV pronto ({linhas:,}). Convertendo XLSB...")

        # Salvar dados texto num parquet temporario pra corrigir depois
        text_parquet_tmp = None
        text_info_v = []
        if text_chunks and text_col_names:
            import tempfile
            df_text = pl.concat(text_chunks, how="vertical_relaxed")
            text_parquet_tmp = Path(tempfile.gettempdir()) / f"_tmp_text_{os.getpid()}.parquet"
            df_text.write_parquet(str(text_parquet_tmp))
            del df_text

            with open(tmp_csv, "r", encoding="utf-8-sig") as hf:
                csv_header = hf.readline().strip().split(";")
            for col_name in text_col_names:
                if col_name in csv_header:
                    text_info_v.append((col_name, csv_header.index(col_name)))
        del text_chunks

        try:
            _csv_para_xlsb_rapido(str(tmp_csv), str(xlsb_out),
                                  parquet_path=str(text_parquet_tmp) if text_parquet_tmp else None,
                                  text_col_info=text_info_v if text_info_v else None,
                                  progress_callback=progress_callback)
        finally:
            try:
                tmp_csv.unlink()
            except Exception:
                pass
            if text_parquet_tmp:
                try:
                    text_parquet_tmp.unlink()
                except Exception:
                    pass

        return linhas, 1

    if tipo == "ENTRADA":
        excluir_cfop = ["1923", "2923", "1915", "2915", "1154", "2154", "1403", "2403", "1555", "2555"]

        ordem_entrada = [
            "ID_ORIGEM", "EMPRESA", "FILIAL", "Divisão", "UF", "DTEMIS", "DTENTR",
            "INFEM_NUM", "CFOP_COD", "VAL_ICMS", "Mapeamento", "TRIBICM", "IND_CANC",
            "VALSUBST_ICMS", "IND_TRIB_SUBSTRIB", "VAL_IPI", "TRIBIPI", "VAL_CONT",
            "ALIQ_DIFICMS", "BAS_ICMS", "ALIQ_ICMS", "ISENTA_ICMS", "VAL_REDICMS",
            "OUTRA_ICMS", "BASSUBST_ICMS", "ALIQ_ST", "VAL_ISEN_SUBSTRIB",
            "VAL_OUTR_SUBSTRIB", "BAS_IPI", "ALIQ_IPI", "ISENTA_IPI", "VAL_REDIPI",
            "OUTRA_IPI", "NOPE_COD", "VAR02", "CNPJ/CPF", "IE", "MOD_DOC", "TDOC_COD",
            "SERIE", "CATG_COD", "CADG_COD", "NUM_ITEM", "MATE_COD", "DSC", "CCUS_COD",
            "CNBM_COD", "UNID_COD_PADRAO", "QTD", "PES_LIQ", "VAL_PRECOUNIT",
            "VAL_PRECOTOT", "VAL_DESC", "VAL_FRETE", "VAL_SEGUR", "IND_MOV",
            "COD_CONT", "NUM_FCI", "FCP_DEST", "ICMS_DEST", "ICMS_ORIG", "FCP_ST",
            "SIST_ORIGEM", "USUA_ORIGEM", "DATA_CRIACAO", "CHAVE DA NOTA", "Nome do Arquivo",
        ]

        mapa_entrada = {
            "ID_ORIGEM": "ID Origem",
            "EMPRESA": "Empresa",
            "FILIAL": "Filial",
            "Divisão": "Divisão",
            "UF": "Unidade Federativa",
            "DTEMIS": "Emissão",
            "DTENTR": "Entrada",
            "INFEM_NUM": "Nota Fiscal",
            "CFOP_COD": "CFOP",
            "VAL_ICMS": "Valor do ICMS",
            "Mapeamento": "Mapeamento",
            "TRIBICM": "Indicador de Tributação de ICMS",
            "IND_CANC": "Indicador de Cancelamento",
            "VALSUBST_ICMS": "Valor Subst. ICMS",
            "IND_TRIB_SUBSTRIB": "Tributação.Subs.Trib.",
            "VAL_IPI": "Valor do IPI",
            "TRIBIPI": "Indicador Trib. IPI",
            "VAL_CONT": "Valor Contábil",
            "ALIQ_DIFICMS": "Alíq. Dif. ICMS",
            "BAS_ICMS": "Base de ICMS",
            "ALIQ_ICMS": "Alíquota de ICMS",
            "ISENTA_ICMS": "Valor Isenta ICMS",
            "VAL_REDICMS": "Valor de Redução da Base",
            "OUTRA_ICMS": "Valor Outras ICMS",
            "BASSUBST_ICMS": "Base Subst. ICMS",
            "ALIQ_ST": "Alíquota do ICMS - ST",
            "VAL_ISEN_SUBSTRIB": "Isent. Subst.Trib.",
            "VAL_OUTR_SUBSTRIB": "Outr. Subst.Trib.",
            "BAS_IPI": "Valor de Base de IPI",
            "ALIQ_IPI": "Aliq. de IPI",
            "ISENTA_IPI": "Valor Isenta IPI",
            "VAL_REDIPI": "Valor de Redução do IPI",
            "OUTRA_IPI": "Valor Outras IPI",
            "NOPE_COD": "Natureza da Operação",
            "VAR02": "OpenFlex 02",
            "CNPJ/CPF": "CNPJ/CPF",
            "IE": "Inscricao Estadual",
            "MOD_DOC": "Modelo de Documento",
            "TDOC_COD": "Tipo de Documento",
            "SERIE": "Serie",
            "CATG_COD": "Categoria",
            "CADG_COD": "Código PF/PJ",
            "NUM_ITEM": "Numero do Item",
            "MATE_COD": "Material",
            "DSC": "Descrição Complementar",
            "CCUS_COD": "Centro de Custo",
            "CNBM_COD": "Classificação Fiscal",
            "UNID_COD_PADRAO": "Unid. Padrão de Venda",
            "QTD": "Quantidade",
            "PES_LIQ": "Peso Liquido",
            "VAL_PRECOUNIT": "Preco Unitario",
            "VAL_PRECOTOT": "Preco Total",
            "VAL_DESC": "Desconto",
            "VAL_FRETE": "Frete",
            "VAL_SEGUR": "Seguro",
            "IND_MOV": "Indicador de Movimento",
            "COD_CONT": "Código Contábil",
            "NUM_FCI": "Número da FCI",
            "FCP_DEST": "Valor FCP UF Destino",
            "ICMS_DEST": "Valor ICMS UF Destino",
            "ICMS_ORIG": "Valor ICMS UF Origem",
            "FCP_ST": "Valor Unitário do FCP ST convertido",
            "SIST_ORIGEM": "Sistema de Origem",
            "USUA_ORIGEM": "Usuário de Origem",
            "DATA_CRIACAO": "Data Integração",
            "CHAVE DA NOTA": "Chave Nota Fiscal",
            "Nome do Arquivo": "Nome Arquivo",
        }

        def processar_batch_entrada(df):
            df = normalizar_colunas_filtro(df, ["CFOP_COD", "IND_CANC", "TRIBICMS", "TRIBICM"])
            col_trib = "TRIBICMS" if "TRIBICMS" in df.columns else ("TRIBICM" if "TRIBICM" in df.columns else None)

            filtros = []
            if col_trib is not None:
                filtros.append(pl.col(col_trib) == "S")
            if "IND_CANC" in df.columns:
                filtros.append(pl.col("IND_CANC") == "N")
            if "CFOP_COD" in df.columns:
                filtros.append(~pl.col("CFOP_COD").is_in(excluir_cfop))

            if filtros:
                expr = filtros[0]
                for f in filtros[1:]:
                    expr = expr & f
                df = df.filter(expr)

            df = selecionar_colunas_existentes(df, ordem_entrada)
            df = renomear_colunas_parcial(df, mapa_entrada)
            return df

        periodo = _extrair_periodo_do_parquet(parquet_path)
        sufixo_periodo = f"_{periodo}" if periodo else ""
        nome_csv = f"BASE_VIVO_ENTRADA{sufixo_periodo}.xlsb"
        processar_batch = processar_batch_entrada

    elif tipo == "SAIDA":
        ordem_saida = lista_unica([
            "Índice", "Fonte", "Período", "ID_ORIGEM", "EMPRESA", "FILIAL", "Divisão", "UF",
            "INFSM_DTEM", "INFSM_NUM", "CFOP_COD", "INFSM_VAL_ICMS", "Mapeamento", "I_2",
            "IND_CANC", "INFSM_VALSUBST_ICMS", "IND_TRIB_SUBSTRIB", "INFSM_VAL_IPI", "I",
            "INFSM_VAL_CONT", "INFSM_ALIQ_DIFICMS", "INFSM_BAS_ICMS", "INFSM_ALIQ_ICMS",
            "INFSM_ISENTA_ICMS", "INFSM_VAL_REDICMS", "INFSM_OUTRA_ICMS",
            "INFSM_BASSUBST_ICMS", "INFSM_ALIQ_ST", "VAL_ISEN_SUBSTRIB", "VAL_OUTR_SUBSTRIB",
            "INFSM_BAS_IPI", "INFSM_ALIQ_IPI", "INFSM_ISENTA_IPI", "INFSM_VAL_REDIPI",
            "INFSM_OUTRA_IPI", "NOPE_COD", "VAR02", "CNPJ/CPF", "IE", "MOD_DOC", "TDOC_COD",
            "INFSM", "CATG_COD", "CADG_COD", "INFSM_2", "MATE_COD", "INFSM_DSC", "CCUS_COD",
            "CNBM_COD", "UNID_COD_VENDA", "INFSM_QTD", "INFSM_PES_LIQ", "INFSM_VAL_PRECOUNIT",
            "INFSM_VAL_PRECOTOT", "INFSM_VAL_DESC", "INFSM_VAL_FRETE", "INFSM_VAL_SEGUR",
            "INFSM_COD_CONT", "INFSM_NUM_FCI", "INFSM_FCP_DEST", "INFSM_ICMS_DEST",
            "INFSM_ICMS_ORIG", "INFSM_FCP_ST_REST", "SIST_ORIGEM", "USUA_ORIGEM",
            "DATA_CRIACAO", "MNFSM_CHV_NFE", "Nome do Arquivo", "INFSM_VAL_DESP", "FEDE_COD",
            "INFSM_ICMS_FRETE", "INFSM_CHASSI", "LIPI_COD", "INFSM_FCP_PRO", "INFSM_FCP_RET",
            "INFSM_FCP_ST", "INFSM_QTD_CONV", "INFSM_VLR_CONV", "INFSM_ICMS_CONV",
            "INFSM_VAL_ICMS_OP_CONV", "INFSM_BC_ICMS_ST_CONV", "INFSM_ICMS_ST_EST",
            "INFSM_FCP_ST_EST", "INFSM_ICMS_ST_REST", "INFSM_ICMS_ST_COMP",
            "INFSM_FCP_ST_COMP", "TMRC_COD",
        ])

        mapa_saida = {
            "Índice": "Índice",
            "Fonte": "Fonte",
            "Período": "Período",
            "ID_ORIGEM": "ID Origem",
            "EMPRESA": "Empresa",
            "FILIAL": "Filial",
            "Divisão": "Divisão",
            "UF": "UF",
            "INFSM_DTEM": "Data Emissão",
            "INFSM_NUM": "Nota Fiscal",
            "CFOP_COD": "CFOP",
            "INFSM_VAL_ICMS": "Vr. de ICMS",
            "Mapeamento": "Mapeamento",
            "I_2": "Indicador de Trib. ICMS",
            "IND_CANC": "Indicador de Cancelamento",
            "INFSM_VALSUBST_ICMS": "Vr. do ICMS por Substituto",
            "IND_TRIB_SUBSTRIB": "Indicador deTrib. Subs.Trib.",
            "INFSM_VAL_IPI": "Vr. do IPI",
            "I": "Indicador de Trib. IPI",
            "INFSM_VAL_CONT": "Vlr. Contábil",
            "INFSM_ALIQ_DIFICMS": "Alíquota ICMS DIFAL",
            "INFSM_BAS_ICMS": "Base de ICMS",
            "INFSM_ALIQ_ICMS": "Aliquota ICMS",
            "INFSM_ISENTA_ICMS": "Isenta ICMS",
            "INFSM_VAL_REDICMS": "Vr. de Redução de ICMS",
            "INFSM_OUTRA_ICMS": "Vlr. outra ICMS",
            "INFSM_BASSUBST_ICMS": "Vr. Base de Substituição ICMS",
            "INFSM_ALIQ_ST": "Aliquota do ICMS-ST",
            "VAL_ISEN_SUBSTRIB": "Valor Isent. Subst.Trib.",
            "VAL_OUTR_SUBSTRIB": "Valor Outr. Subst.Trib",
            "INFSM_BAS_IPI": "Base de IPI",
            "INFSM_ALIQ_IPI": "Aliquota de IPI",
            "INFSM_ISENTA_IPI": "Isenta de IPI",
            "INFSM_VAL_REDIPI": "Valor de Redução do IPI",
            "INFSM_OUTRA_IPI": "Vlr. de Outra de IPI",
            "NOPE_COD": "Natureza",
            "VAR02": "OpenFlex 02",
            "CNPJ/CPF": "CNPJ/CPF",
            "IE": "Inscrição Estadual",
            "MOD_DOC": "Modelo de Documento",
            "TDOC_COD": "Tipo de Documento",
            "INFSM": "Série",
            "CATG_COD": "Categoria",
            "CADG_COD": "Código PF\\PJ",
            "INFSM_2": "Item",
            "MATE_COD": "Material",
            "INFSM_DSC": "Descrição",
            "CCUS_COD": "Centro de Custo",
            "CNBM_COD": "NCM",
            "UNID_COD_VENDA": "Unidade Código Padrão",
            "INFSM_QTD": "Quantidade",
            "INFSM_PES_LIQ": "Peso Líquido",
            "INFSM_VAL_PRECOUNIT": "Preço Unitário",
            "INFSM_VAL_PRECOTOT": "Preço Total",
            "INFSM_VAL_DESC": "Valor de Desconto",
            "INFSM_VAL_FRETE": "Vr. do Frete",
            "INFSM_VAL_SEGUR": "Vr. Seguro",
            "INFSM_COD_CONT": "Código Contábil",
            "INFSM_NUM_FCI": "Número da FCI",
            "INFSM_FCP_DEST": "Valor FCP UF Destino",
            "INFSM_ICMS_DEST": "Valor ICMS UF Destino",
            "INFSM_ICMS_ORIG": "Valor ICMS UF Origem",
            "INFSM_FCP_ST_REST": "Valor Unitário FCP ST Convertido A Restituir",
            "SIST_ORIGEM": "Sistema de Origem",
            "USUA_ORIGEM": "Usuário de Origem",
            "DATA_CRIACAO": "Data de Criação",
            "MNFSM_CHV_NFE": "Chave de Acesso",
            "Nome do Arquivo": "FileName",
            "INFSM_VAL_DESP": "Vr. Despesa",
            "FEDE_COD": "Sit. Trib. Federal",
            "INFSM_ICMS_FRETE": "ICMS sobre Frete",
            "INFSM_CHASSI": "Número do Chassi",
            "LIPI_COD": "CST IPI",
            "INFSM_FCP_PRO": "Valor do Fundo de Combate à Pobreza (FCP) Próprio",
            "INFSM_FCP_RET": "Valor do Fundo de Combate à Pobreza (FCP) Retido",
            "INFSM_FCP_ST": "Valor do Fundo de Combate à Pobreza (FCP) ST",
            "INFSM_QTD_CONV": "Quantidade do Item convertida",
            "INFSM_VLR_CONV": "Valor unitário convertido",
            "INFSM_ICMS_CONV": "Valor unitário ICMS Operação convertido",
            "INFSM_VAL_ICMS_OP_CONV": "Valor unitário ICMS operação de Entrada Convertido",
            "INFSM_BC_ICMS_ST_CONV": "Valor unitário da Base de Cálculo do ICMS ST Convertido",
            "INFSM_ICMS_ST_EST": "Valor Unitário ICMS ST Estoque Convertido",
            "INFSM_FCP_ST_EST": "Valor Unitário FCP ICMS ST Estoque Convertidos",
            "INFSM_ICMS_ST_REST": "Valor Unitario ICMS ST Convertido a Restituir",
            "INFSM_ICMS_ST_COMP": "Valor unitário do ICMS ST Convertido",
            "INFSM_FCP_ST_COMP": "Valor Unitário do FCP ST convertido",
            "TMRC_COD": "Tab105",
        }

        def processar_batch_saida(df):
            df = normalizar_colunas_filtro(df, ["I_2", "IND_CANC"])

            filtros = []
            if "I_2" in df.columns:
                filtros.append(pl.col("I_2") == "S")
            if "IND_CANC" in df.columns:
                filtros.append(pl.col("IND_CANC") == "N")

            if filtros:
                expr = filtros[0]
                for f in filtros[1:]:
                    expr = expr & f
                df = df.filter(expr)

            df = selecionar_colunas_existentes(df, ordem_saida)
            df = renomear_colunas_parcial(df, mapa_saida)
            return df

        periodo = _extrair_periodo_do_parquet(parquet_path)
        sufixo_periodo = f"_{periodo}" if periodo else ""
        nome_csv = f"BASE_VIVO_SAIDA{sufixo_periodo}.xlsb"
        processar_batch = processar_batch_saida

    else:
        raise ValueError("Tipo de movimento não identificado para exportação VIVO.")

    out_xlsx = Path(pasta_destino) / nome_csv

    linhas_gravadas, total_batches = escrever_xlsb_csv_com(
        parquet_path=parquet_path,
        xlsb_out=str(out_xlsx),
        processar_batch=processar_batch
    )

    if progress_callback:
        progress_callback(
            "finalizado_xlsx",
            total_batches,
            total_batches,
            f"Exportação XLSX concluída | gravadas: {linhas_gravadas:,}"
        )

    return out_xlsx


def criar_txt_limpo(path_txt, tmp_txt):
    header_line, header_raw = detectar_header(path_txt)
    if header_line is None:
        return None, None, 0, None

    header = limpar_nomes_colunas(header_raw)
    ncols = len(header)
    idx_dsc = descobrir_idx_dsc(header)
    idx_seguro = descobrir_idx_merge_seguro(header)

    # Escolhe onde absorver os `|` extras. Prioriza uma coluna tardia de
    # free-text (VAR05 etc); só cai no DSC se nenhuma existir no header.
    idx_merge = idx_seguro if idx_seguro is not None else idx_dsc
    kept = 0

    # Buffers maiores reduzem o número de syscalls de write/read em ~100x
    # para esses TXTs grandes. Default do Python é 8KB; 1MB é mais apropriado.
    _BUF = 1024 * 1024

    with open(path_txt, "r", encoding="latin-1", errors="ignore", buffering=_BUF) as fin, \
         open(tmp_txt, "w", encoding="latin-1", errors="ignore", newline="", buffering=_BUF) as fout:

        for _ in range(header_line + 1):
            next(fin)

        # Acumula as linhas corrigidas numa lista e escreve em blocos grandes
        # com writelines(), que é mais eficiente que chamar write() por linha.
        # A lista é reciclada em chunks de 10k para manter o pico de memória
        # baixo em arquivos muito longos.
        buf_out = ["|".join(header) + "\n"]
        FLUSH_EVERY = 10_000

        for linha in fin:
            if linha_eh_lixo(linha):
                continue

            buf_out.append(corrigir_pipe_na_descricao(linha, ncols, idx_merge) + "\n")
            kept += 1

            if len(buf_out) >= FLUSH_EVERY:
                fout.writelines(buf_out)
                buf_out.clear()

        if buf_out:
            fout.writelines(buf_out)

    return header, header_raw, kept, idx_merge


def processar_arquivo(args):
    if len(args) == 3:
        path_txt_str, tmp_dir_str, tipo_mov = args
    else:
        path_txt_str, tmp_dir_str = args
        tipo_mov = None

    path_txt = Path(path_txt_str)
    tmp_dir = Path(tmp_dir_str)

    shard_out = tmp_dir / f"{path_txt.stem}.parquet"
    txt_limpo = tmp_dir / f"{path_txt.stem}__limpo.txt"

    header, header_raw, kept, idx_merge = criar_txt_limpo(path_txt, txt_limpo)
    if header is None:
        return {"arquivo": path_txt.name, "linhas": 0, "ok": False, "motivo": "header não encontrado"}

    df = pl.read_csv(
        txt_limpo,
        separator="|",
        has_header=True,
        encoding="latin1",
        infer_schema_length=0,
        schema_overrides={c: pl.Utf8 for c in header},
        ignore_errors=False,
        truncate_ragged_lines=True,
        quote_char=None,
    )

    # O PIPE_PLACEHOLDER foi gravado na coluna escolhida por `criar_txt_limpo`
    # (que pode ser DSC ou VAR05, conforme o caso). Precisamos decodificá-lo
    # EXATAMENTE naquela coluna.
    if idx_merge is not None:
        col_merge = header[idx_merge]
        df = df.with_columns(
            pl.col(col_merge)
            .str.replace_all(PIPE_PLACEHOLDER, "|")
            .alias(col_merge)
        )

    nome_arquivo = path_txt.name
    divisao_arquivo = extrair_divisao_arquivo(nome_arquivo)

    div_df = carregar_divisoes_df()
    cfop_df = carregar_cfop_df()

    if "DTENTR" in df.columns:
        expr_periodo = (
            pl.when(pl.col("DTENTR").is_not_null() & (pl.col("DTENTR").str.len_chars() >= 10))
            .then(pl.col("DTENTR").str.slice(6, 4) + pl.lit("_") + pl.col("DTENTR").str.slice(3, 2))
            .otherwise(pl.lit(""))
            .alias("Período")
        )
    elif "INFSM_DTEM" in df.columns:
        expr_periodo = (
            pl.when(pl.col("INFSM_DTEM").is_not_null() & (pl.col("INFSM_DTEM").str.len_chars() >= 10))
            .then(pl.col("INFSM_DTEM").str.slice(6, 4) + pl.lit("_") + pl.col("INFSM_DTEM").str.slice(3, 2))
            .otherwise(pl.lit(""))
            .alias("Período")
        )
    else:
        expr_periodo = pl.lit("").alias("Período")

    # Fonte baseada no tipo_movimento selecionado pelo usuario
    if tipo_mov and tipo_mov.upper() == "ENTRADA":
        fonte_label = "Livro de Entrada"
    elif tipo_mov and tipo_mov.upper() == "SAIDA":
        fonte_label = "Livro de Saída"
    else:
        fonte_label = "Livro de Entrada" if "DTENTR" in df.columns else "Livro de Saída"

    df = df.with_columns([
        pl.lit(fonte_label).alias("Fonte"),
        pl.lit(nome_arquivo).alias("Nome do Arquivo"),
        pl.lit(divisao_arquivo).alias("DivArquivo"),
        expr_periodo,
        pl.col("FILIAL").cast(pl.Utf8).str.strip_chars().alias("FILIAL"),
        pl.col("CFOP_COD").cast(pl.Utf8).str.strip_chars().alias("CFOP_COD"),
    ])

    df = df.join(
        div_df.with_columns(
            pl.col("Local de Negócios").cast(pl.Utf8).str.strip_chars().alias("Local de Negócios")
        ),
        left_on="FILIAL",
        right_on="Local de Negócios",
        how="left"
    )

    df = df.join(
        cfop_df,
        left_on="CFOP_COD",
        right_on="CFOP",
        how="left"
    )

    # garantir coluna Mapeamento
    if "Mapeamento" not in df.columns:
        df = df.with_columns(pl.lit("").alias("Mapeamento"))
    else:
        df = df.with_columns(
            pl.col("Mapeamento").fill_null("").alias("Mapeamento")
        )

    df = df.with_columns([
        pl.when(pl.col("FILIAL") == "3007")
        .then(pl.lit("31SC"))
        .when(pl.col("DivArquivo").str.to_uppercase() == "85MN")
        .then(pl.lit("85MG"))
        .when(pl.col("DivArquivo") == "")
        .then(pl.col("Divisão_DePara").fill_null(""))
        .when(pl.col("Divisão_DePara").fill_null("").str.to_uppercase() == pl.col("DivArquivo").str.to_uppercase())
        .then(pl.col("DivArquivo"))
        .otherwise(pl.col("DivArquivo"))
        .alias("Divisão")
    ])

    df = df.with_columns([
        pl.col(c).str.strip_chars().alias(c)
        for c in df.columns
        if df.schema[c] == pl.Utf8
    ])

    ordem = montar_ordem_final(df.columns)
    df = df.select(ordem)

    df = df.with_columns(pl.lit(path_txt.name).alias("__ordem__"))

    df.write_parquet(
        shard_out,
        compression=COMPRESSION,
        row_group_size=ROW_GROUP
    )

    try:
        txt_limpo.unlink(missing_ok=True)
    except Exception:
        pass

    return {"arquivo": path_txt.name, "linhas": df.height, "ok": True, "shard": str(shard_out)}


def consolidar_final(base_dir_str, progress_callback=None):
    t0 = time.time()

    base_dir = Path(base_dir_str)
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    tmp_dir = CACHE_DIR / TMP_NOME

    arquivos = sorted(base_dir.rglob("*.txt"))
    if not arquivos:
        raise FileNotFoundError("Nenhum TXT encontrado na pasta selecionada.")

    tipo_movimento = detectar_tipo_movimento(arquivos)

    tmp_dir.mkdir(parents=True, exist_ok=True)

    if progress_callback:
        progress_callback("preparando", 0, len(arquivos), "Preparando processamento...")

    shard_paths = []
    total_linhas = 0
    total_arquivos = len(arquivos)

    args = [(str(a), str(tmp_dir), tipo_movimento) for a in arquivos]

    for i, arg in enumerate(args, start=1):
        res = processar_arquivo(arg)

        if progress_callback:
            progress_callback(
                "processando_txt",
                i,
                total_arquivos,
                res.get("arquivo", "")
            )

        if not res["ok"]:
            continue

        shard_paths.append(res["shard"])
        total_linhas += res["linhas"]

    if not shard_paths:
        raise ValueError("Nenhum shard gerado.")

    if progress_callback:
        progress_callback("consolidando", 1, 1, "Consolidando shards...")

    df = (
        pl.scan_parquet(shard_paths)
        .with_row_index("Índice", offset=1)
        .collect()
    )

    cols = [c for c in df.columns if c not in {"Índice", "__ordem__"}]
    ordem_final = ["Índice"] + cols
    df = df.select(ordem_final)

    periodos = sorted(set(
        str(x).strip() for x in df["Período"].to_list()
        if str(x).strip()
    ))

    if len(periodos) == 1:
        periodo_txt = periodos[0]
    elif len(periodos) > 1:
        periodo_txt = f"MULTI_{len(periodos)}_PERIODOS"
    else:
        periodo_txt = "SEM_PERIODO"

    parquet_final = CACHE_DIR / f"BASE_INTERNA__{periodo_txt}__{tipo_movimento}.parquet"

    df.write_parquet(parquet_final, compression=COMPRESSION, row_group_size=ROW_GROUP)

    gerar_bases_conferencia_cache(parquet_final, tipo_movimento, base_dir)

    if progress_callback:
        progress_callback("finalizado", 1, 1, parquet_final.name)

    return parquet_final, df.height, round(time.time() - t0, 2), tipo_movimento