#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import re
import sys
import time
import unicodedata
import json
from pathlib import Path

import pdfplumber
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle


# =========================================================
# CONFIG GERAL
# =========================================================
ARQ_DIVISAO = "Tabela_Divisao.csv"
ARQ_CFOP_MAP = "Mapeamento_CFOP.csv"

TITULO_ALVO = "LIVRO REGISTRO DE APURAÇÃO DO ICMS - RAICMS - MODELO P"
TITULO_RESUMO = "RESUMO DA APURAÇÃO DO IMPOSTO".upper()
CACHE_DIR = Path.home() / "AppData" / "Local" / "ValidadorVIVO"
CACHE_RAICMS_META = CACHE_DIR / "raicms_meta.json"


# =========================================================
# RECURSOS
# =========================================================
def caminho_recurso(nome_arquivo):
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, nome_arquivo)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), nome_arquivo)


# =========================================================
# PROCESSAMENTO RAICMS
# =========================================================
def normalizar_espacos(txt: str) -> str:
    return re.sub(r"\s+", " ", txt or "").strip()


def normalizar_texto(txt: str) -> str:
    txt = txt or ""
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    txt = normalizar_espacos(txt).upper()
    return txt


def converter_numero_br(valor):
    if valor is None:
        return None

    valor = str(valor).strip()
    if not valor:
        return None

    valor = valor.replace(".", "").replace(",", ".")
    try:
        return float(valor)
    except Exception:
        return None


def formatar_periodo(periodo):
    if not periodo:
        return ""
    return str(periodo).replace("/", ".")

def ler_txt_com_fallback(caminho_txt):
    encodings = ["utf-8-sig", "utf-8", "latin1"]

    for enc in encodings:
        try:
            with open(caminho_txt, "r", encoding=enc, errors="strict") as f:
                return f.read()
        except Exception:
            continue

    with open(caminho_txt, "r", encoding="latin1", errors="ignore") as f:
        return f.read()

def extrair_cabecalho_pagina(texto):
    firma = ""
    ie = ""
    cnpj = ""
    periodo = ""
    folha = ""

    m_firma = re.search(r"FIRMA\s*:\s*(.+)", texto)
    if m_firma:
        firma = normalizar_espacos(m_firma.group(1))

    m_ie = re.search(r"INSCR\.\s*EST\.\s*:\s*([^\n]+?)\s+CNPJ\s*:\s*([0-9\.\-\/]+)", texto)
    if m_ie:
        ie = normalizar_espacos(m_ie.group(1))
        cnpj = normalizar_espacos(m_ie.group(2))
    else:
        m_ie2 = re.search(r"INSCR\.\s*EST\.\s*:\s*(.+)", texto)
        if m_ie2:
            ie = normalizar_espacos(m_ie2.group(1))

        m_cnpj = re.search(r"CNPJ\s*:\s*([0-9\.\-\/]+)", texto)
        if m_cnpj:
            cnpj = normalizar_espacos(m_cnpj.group(1))

    m_periodo = re.search(r"Per[ií]odo\s*:\s*(\d{2}/\d{4})", texto, flags=re.IGNORECASE)
    if m_periodo:
        periodo = m_periodo.group(1)

    m_folha = re.search(r"FOLHA\s*:\s*([^\n]+)", texto)
    if m_folha:
        folha = normalizar_espacos(m_folha.group(1))

    return {
        "Firma": firma,
        "IE": ie,
        "CNPJ": cnpj,
        "Período": periodo,
        "Folha": folha
    }


def extrair_filial_do_arquivo(caminho_pdf):
    nome = os.path.splitext(os.path.basename(caminho_pdf))[0]
    nome_up = nome.upper()

    if "PAY" in nome_up:
        return "PAY"

    # PRIORIDADE MÁXIMA: FL3164 / FL 3164 / FL-3164 / FL_3164
    m = re.search(r"\bFL[\s_\-]*(\d{4})\b", nome_up)
    if m:
        return m.group(1)

    # Se não achou FL, tenta outros blocos de 4 dígitos no nome do PDF,
    # mas ignorando os que forem claramente período (MMYY ou YYYY)
    candidatos_pdf = re.findall(r"(?<!\d)(\d{4})(?!\d)", nome_up)

    candidatos_pdf_filtrados = []
    for c in candidatos_pdf:
        if re.fullmatch(r"20\d{2}", c):
            continue

        mes = int(c[:2])
        ano2 = c[2:]
        if 1 <= mes <= 12 and re.fullmatch(r"\d{2}", ano2):
            continue

        candidatos_pdf_filtrados.append(c)

    if candidatos_pdf_filtrados:
        return candidatos_pdf_filtrados[0]

    return ""

def encontrar_filial_por_txt(caminho_pdf, cnpj, ie):
    pasta = os.path.dirname(caminho_pdf)

    try:
        txts = [f for f in os.listdir(pasta) if f.lower().endswith(".txt")]
    except Exception:
        return ""

    candidatos = []

    cnpj_norm = re.sub(r"\D", "", cnpj or "")
    ie_norm = re.sub(r"\D", "", ie or "")

    for txt in txts:
        caminho_txt = os.path.join(pasta, txt)

        try:
            conteudo = ler_txt_com_fallback(caminho_txt)
        except Exception:
            continue

        conteudo_cnpj_ie = conteudo
        conteudo_cnpj_ie_norm = re.sub(r"\D", "", conteudo_cnpj_ie)

        if cnpj_norm and cnpj_norm not in conteudo_cnpj_ie_norm:
            continue

        if ie_norm and ie_norm not in conteudo_cnpj_ie_norm:
            continue

        nome_txt = os.path.splitext(os.path.basename(txt))[0].upper()

        m = re.search(r"\bFL[\s_\-]*(\d{4})\b", nome_txt)
        if m:
            candidatos.append(m.group(1))
            continue

        nums = re.findall(r"(?<!\d)(\d{4})(?!\d)", nome_txt)

        candidatos_filtrados = []
        for n in nums:
            if re.fullmatch(r"20\d{2}", n):
                continue

            mes = int(n[:2])
            ano2 = n[2:]
            if 1 <= mes <= 12 and re.fullmatch(r"\d{2}", ano2):
                continue

            candidatos_filtrados.append(n)

        if candidatos_filtrados:
            candidatos.append(candidatos_filtrados[0])

    candidatos = list(dict.fromkeys(candidatos))

    if len(candidatos) == 1:
        return candidatos[0]

    return ""

def encontrar_filial_priorizando_txts_da_pasta(caminho_arquivo, cnpj="", ie=""):
    pasta = os.path.dirname(caminho_arquivo)

    try:
        txts = sorted(
            os.path.join(pasta, f)
            for f in os.listdir(pasta)
            if f.lower().endswith(".txt")
        )
    except Exception:
        txts = []

    # 1) Prioridade máxima: tentar extrair filial diretamente do nome dos TXTs da pasta
    for caminho_txt in txts:
        filial_txt = extrair_filial_do_arquivo(caminho_txt)
        if filial_txt:
            return filial_txt

    # 2) Se não achou pelo nome, tenta pelo conteúdo/CNPJ/IE
    if txts:
        filial_por_txt = encontrar_filial_por_txt(caminho_arquivo, cnpj, ie)
        if filial_por_txt:
            return filial_por_txt

    return ""

def carregar_mapa_divisao():
    arq = caminho_recurso(ARQ_DIVISAO)

    tentativas = [
        {"sep": None, "engine": "python", "encoding": "utf-8-sig"},
        {"sep": ";", "encoding": "utf-8-sig"},
        {"sep": ",", "encoding": "utf-8-sig"},
        {"sep": None, "engine": "python", "encoding": "latin1"},
        {"sep": ";", "encoding": "latin1"},
        {"sep": ",", "encoding": "latin1"},
    ]

    ultimo_erro = None
    df = None

    for cfg in tentativas:
        try:
            df = pd.read_csv(arq, **cfg)
            break
        except Exception as e:
            ultimo_erro = e

    if df is None:
        raise RuntimeError(f"Erro ao ler Tabela_Divisao.csv: {ultimo_erro}")

    cols_norm = {normalizar_texto(c): c for c in df.columns}

    col_filial = cols_norm.get(normalizar_texto("Local de Negócios"))
    col_divisao = cols_norm.get(normalizar_texto("Divisão"))

    if not col_filial or not col_divisao:
        raise RuntimeError(
            "Não encontrei as colunas 'Local de Negócios' e/ou 'Divisão' no Tabela_Divisao.csv"
        )

    mapa = {}

    for _, row in df[[col_filial, col_divisao]].dropna(how="all").iterrows():
        filial = normalizar_espacos(str(row[col_filial])) if pd.notna(row[col_filial]) else ""
        divisao = normalizar_espacos(str(row[col_divisao])) if pd.notna(row[col_divisao]) else ""

        filial_limpa = re.sub(r"\.0$", "", filial.strip())
        if not filial_limpa or not divisao:
            continue

        chaves = {filial_limpa.upper()}

        if re.fullmatch(r"\d+", filial_limpa):
            chaves.add(filial_limpa.zfill(4))

        for chave in chaves:
            mapa.setdefault(chave, [])
            if divisao not in mapa[chave]:
                mapa[chave].append(divisao)

    return mapa

def resolver_divisao_por_filial_e_pasta(caminho_arquivo, filial, mapa_divisao):
    filial_norm = (filial or "").strip().upper()

    if not filial_norm:
        return ""

    if filial_norm == "PAY":
        return "PAY"

    divisoes = mapa_divisao.get(filial_norm, [])

    def normalizar_divisao_final(div):
        div = (div or "").strip().upper()
        if div == "85MN":
            return "85MG"
        return div

    if not divisoes:
        pasta_nome = os.path.basename(os.path.dirname(caminho_arquivo)).strip().upper()
        return normalizar_divisao_final(pasta_nome)

    if len(divisoes) == 1:
        return normalizar_divisao_final(divisoes[0])

    pasta_nome = os.path.basename(os.path.dirname(caminho_arquivo)).strip().upper()
    pasta_nome = normalizar_divisao_final(pasta_nome)

    for divisao in divisoes:
        if normalizar_divisao_final(divisao) == pasta_nome:
            return normalizar_divisao_final(divisao)

    return normalizar_divisao_final(divisoes[0])

def carregar_mapa_cfop():
    arq = caminho_recurso(ARQ_CFOP_MAP)

    tentativas = [
        {"sep": ";", "encoding": "utf-8-sig"},
        {"sep": ",", "encoding": "utf-8-sig"},
        {"sep": ";", "encoding": "latin1"},
        {"sep": ",", "encoding": "latin1"},
    ]

    ultimo_erro = None
    df = None

    for cfg in tentativas:
        try:
            df = pd.read_csv(arq, **cfg)
            break
        except Exception as e:
            ultimo_erro = e

    if df is None:
        raise RuntimeError(f"Erro ao ler Mapeamento_CFOP.csv: {ultimo_erro}")

    cols_norm = {normalizar_texto(c): c for c in df.columns}
    col_cfop = cols_norm.get(normalizar_texto("CFOP"))
    col_map = cols_norm.get(normalizar_texto("Mapeamento"))

    if not col_cfop or not col_map:
        raise RuntimeError("Não encontrei as colunas 'CFOP' e/ou 'Mapeamento' no Mapeamento_CFOP.csv")

    mapa = {}
    for _, row in df[[col_cfop, col_map]].dropna(how="all").iterrows():
        cfop = normalizar_espacos(str(row[col_cfop])) if pd.notna(row[col_cfop]) else ""
        mapeamento = normalizar_espacos(str(row[col_map])) if pd.notna(row[col_map]) else ""
        if cfop:
            mapa[str(cfop).strip()] = mapeamento

    return mapa


def aplicar_sinal_entrada(valor, tipo):
    if valor is None or pd.isna(valor):
        return valor
    if (tipo or "").strip().upper() == "ENTRADA":
        return -abs(float(valor))
    return abs(float(valor))


def definir_status_por_descricao(descricao):
    desc_n = normalizar_texto(descricao)
    if desc_n == "SALDO CREDOR A TRANSPORTAR" or desc_n == "IMPOSTO A RECOLHER":
        return "ICMS a Recolher ou Recuperar"
    return "Entrada / Saída"


# =========================================================
# CFOP
# =========================================================
def eh_linha_numerica_tabela(linha):
    linha = linha.strip()

    if not re.match(r"^\d{4}\b", linha):
        return False

    nums = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}", linha)
    return len(nums) >= 3


def parse_linha_tabela(linha):
    linha = normalizar_espacos(linha)

    m_cfop = re.match(r"^(\d{4})\b", linha)
    if not m_cfop:
        return None

    cfop = m_cfop.group(1)
    nums = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}", linha)

    if len(nums) < 5:
        return None

    ultimos_5 = nums[-5:]

    return {
        "CFOP": cfop,
        "Valores Contábeis": converter_numero_br(ultimos_5[0]),
        "Base de Cálculo": converter_numero_br(ultimos_5[1]),
        "Imposto Creditado": converter_numero_br(ultimos_5[2]),
        "Isentas ou não Trib.": converter_numero_br(ultimos_5[3]),
        "Outras": converter_numero_br(ultimos_5[4]),
    }


def linha_indica_secao_cfop(linha):
    l = normalizar_texto(linha)

    if l == "ENTRADAS":
        return "Entrada"

    if "SAIDAS" in l:
        return "Saída"

    return None


def deve_ignorar_linha_cfop(linha):
    l = normalizar_espacos(linha).upper()

    ignorar_se_conter = [
        "LIVRO REGISTRO DE APURAÇÃO DO ICMS",
        "REGISTRO DE APURAÇÃO DO ICMS",
        "RESUMO DA APURAÇÃO DO IMPOSTO",
        "ICMS - VALORES FISCAIS",
        "OPERAÇÕES COM CRÉDITO DO IMPOSTO",
        "OPERACOES COM CREDITO DO IMPOSTO",
        "OPERAÇÕES SEM CRÉDITO DO IMPOSTO",
        "OPERACOES SEM CREDITO DO IMPOSTO",
        "CODIFICAÇÃO",
        "CONTÁBIL",
        "CONTABIL",
        "FISCAL",
        "VALORES CONTÁBEIS",
        "VALORES CONTABEIS",
        "BASE DE CÁLCULO",
        "BASE DE CALCULO",
        "IMPOSTO CREDITADO",
        "ISENTAS OU NÃO TRIB.",
        "ISENTAS OU NAO TRIB.",
        "OUTRAS",
        "SUBTOTAIS",
        "TOTAIS",
        "1.00 DO ESTADO",
        "2.00 DE OUTROS ESTADOS",
        "3.00 DO EXTERIOR",
        "FIRMA :",
        "INSCR. EST. :",
        "CNPJ :",
        "FOLHA :",
        "PERÍODO :",
        "PERIODO :"
    ]

    return any(x in l for x in ignorar_se_conter)


def processar_pdf_cfop(caminho_pdf, mapa_divisao, mapa_cfop):
    registros = []
    arquivo_pdf = os.path.basename(caminho_pdf)
    filial = extrair_filial_do_arquivo(caminho_pdf)

    with pdfplumber.open(caminho_pdf) as pdf:
        for num_pagina, pagina in enumerate(pdf.pages, start=1):
            texto = pagina.extract_text() or ""

            if TITULO_ALVO not in texto:
                continue

            if TITULO_RESUMO in texto.upper():
                continue

            meta = extrair_cabecalho_pagina(texto)

            if not filial:
                filial = encontrar_filial_priorizando_txts_da_pasta(
                    caminho_pdf,
                    meta["CNPJ"],
                    meta["IE"]
                )

            divisao = resolver_divisao_por_filial_e_pasta(caminho_pdf, filial, mapa_divisao)

            linhas = texto.splitlines()
            tipo_atual = None

            for linha in linhas:
                linha_limpa = linha.strip()

                if not linha_limpa:
                    continue

                novo_tipo = linha_indica_secao_cfop(linha_limpa)
                if novo_tipo:
                    tipo_atual = novo_tipo
                    continue

                if deve_ignorar_linha_cfop(linha_limpa):
                    continue

                if eh_linha_numerica_tabela(linha_limpa):
                    dados_linha = parse_linha_tabela(linha_limpa)
                    if dados_linha:
                        dados_linha["Imposto Creditado"] = aplicar_sinal_entrada(
                            dados_linha["Imposto Creditado"], tipo_atual
                        )

                        registros.append({
                            "Arquivo PDF": arquivo_pdf,
                            "Filial": filial,
                            "Divisão": divisao,
                            "Página": num_pagina,
                            "Firma": meta["Firma"],
                            "IE": meta["IE"],
                            "CNPJ": meta["CNPJ"],
                            "Período": meta["Período"],
                            "Tipo": tipo_atual,
                            "Status": "Entrada / Saída",
                            "Mapeamento": mapa_cfop.get(str(dados_linha["CFOP"]).strip(), ""),
                            **dados_linha
                        })
    return registros


# =========================================================
# RESUMO APURAÇÃO
# =========================================================
def linha_indica_secao_resumo(linha):
    l = normalizar_texto(linha)

    if "DEBITO DO IMPOSTO" in l:
        return "Débito do Imposto"
    if "CREDITO DO IMPOSTO" in l:
        return "Crédito do Imposto"
    if "APURACAO DO SALDO" in l:
        return "Apuração do Saldo"

    return None


def deve_ignorar_linha_resumo(linha):
    l = normalizar_texto(linha)

    ignorar = [
        "LIVRO REGISTRO DE APURACAO DO ICMS",
        "RAICMS - MODELO P",
        "RESUMO DA APURACAO DO IMPOSTO",
        "VALORES",
        "COLUNA AUXILIAR",
        "SOMAS",
        "FIRMA :",
        "INSCR. EST. :",
        "CNPJ :",
        "FOLHA :",
        "PERIODO :",
        "INFORMACOES COMPLEMENTARES",
        "GUIAS DE RECOLHIMENTO",
        "GUIA DE INFORMACAO",
        "OBSERVACOES",
        "DATA DA ENTREGA",
        "LOCAL DA ENTREGA",
        "ORGAO ARRECADADOR",
        "BANCO :",
        "AGENCIA :"
    ]

    return any(x in l for x in ignorar)


def parse_linha_resumo(linha):
    linha_original = linha.rstrip()
    linha = linha_original.strip()

    if not linha:
        return None

    numero = ""
    descricao = ""

    m_fim_2 = re.search(
        r"\s+(\d{1,3}(?:\.\d{3})*,\d{2})\s+(\d{1,3}(?:\.\d{3})*,\d{2})\s*$",
        linha
    )
    m_fim_1 = re.search(
        r"\s+(\d{1,3}(?:\.\d{3})*,\d{2})\s*$",
        linha
    )

    col_aux = None
    somas = None
    linha_sem_valores = linha

    if m_fim_2:
        col_aux = converter_numero_br(m_fim_2.group(1))
        somas = converter_numero_br(m_fim_2.group(2))
        linha_sem_valores = linha[:m_fim_2.start()].rstrip()
    elif m_fim_1:
        valor_final = converter_numero_br(m_fim_1.group(1))
        linha_sem_valores = linha[:m_fim_1.start()].rstrip()
    else:
        valor_final = None

    m_num = re.match(r"^(\d{3})\s*-\s*(.+?)\s*$", linha_sem_valores)
    if m_num:
        numero = m_num.group(1).strip()
        descricao = m_num.group(2).strip()
        if m_fim_1 and not m_fim_2:
            somas = valor_final
    else:
        m_sem_num = re.match(r"^-\s*(.+?)\s*$", linha_sem_valores)
        if m_sem_num:
            numero = ""
            descricao = m_sem_num.group(1).strip()
            if m_fim_1 and not m_fim_2:
                col_aux = valor_final
        else:
            return None

    return {
        "Número": numero,
        "Descrição": descricao,
        "Coluna Auxiliar": col_aux,
        "Somas": somas
    }


def definir_tipo_resumo(secao, descricao):
    secao_n = normalizar_texto(secao)
    desc_n = normalizar_texto(descricao)

    if "CREDITO" in secao_n:
        return "Entrada"

    if "DEBITO" in secao_n:
        return "Saída"

    if "APURACAO" in secao_n:
        if "DEDUCAO" in desc_n or "DEDUCOES" in desc_n:
            return "Entrada"
        if "IMPOSTO A RECOLHER" in desc_n:
            return "Entrada"
        if "SALDO CREDOR A TRANSPORTAR" in desc_n:
            return "Saída"
        if "SALDO DEVEDOR" in desc_n:
            return "Saída"

    return ""


def processar_pdf_resumo(caminho_pdf, mapa_divisao):
    registros = []
    arquivo_pdf = os.path.basename(caminho_pdf)
    filial = extrair_filial_do_arquivo(caminho_pdf)

    with pdfplumber.open(caminho_pdf) as pdf:
        seq = 0

        for num_pagina, pagina in enumerate(pdf.pages, start=1):
            texto = pagina.extract_text() or ""

            if TITULO_ALVO not in texto:
                continue

            if TITULO_RESUMO not in texto.upper():
                continue

            meta = extrair_cabecalho_pagina(texto)

            if not filial:
                filial = encontrar_filial_priorizando_txts_da_pasta(
                    caminho_pdf,
                    meta["CNPJ"],
                    meta["IE"]
                )

            divisao = resolver_divisao_por_filial_e_pasta(caminho_pdf, filial, mapa_divisao)

            linhas = texto.splitlines()
            secao_atual = None

            for linha in linhas:
                linha_limpa = linha.strip()

                if not linha_limpa:
                    continue

                nova_secao = linha_indica_secao_resumo(linha_limpa)
                if nova_secao:
                    secao_atual = nova_secao
                    continue

                if deve_ignorar_linha_resumo(linha_limpa):
                    continue

                dados = parse_linha_resumo(linha_limpa)
                if dados:
                    seq += 1
                    tipo = definir_tipo_resumo(secao_atual, dados["Descrição"])
                    dados["Somas"] = aplicar_sinal_entrada(dados["Somas"], tipo)
                    status = definir_status_por_descricao(dados["Descrição"])

                    registros.append({
                        "_seq": seq,
                        "Arquivo PDF": arquivo_pdf,
                        "Filial": filial,
                        "Divisão": divisao,
                        "Página": num_pagina,
                        "Firma": meta["Firma"],
                        "IE": meta["IE"],
                        "CNPJ": meta["CNPJ"],
                        "Período": meta["Período"],
                        "Seção": secao_atual,
                        "Tipo": tipo,
                        "Status": status,
                        **dados
                    })

    return registros


def processar_txt_cfop(caminho_txt, mapa_divisao, mapa_cfop):
    registros = []
    arquivo_txt = os.path.basename(caminho_txt)
    filial = extrair_filial_do_arquivo(caminho_txt)

    try:
        texto_total = ler_txt_com_fallback(caminho_txt)
    except Exception:
        return registros

    paginas = re.split(
        r"(?=LIVRO REGISTRO DE APURAÇÃO DO ICMS - RAICMS - MODELO P9|LIVRO REGISTRO DE APURAÇÃO DO ICMS - RAICMS - MODELO P)",
        texto_total
    )

    num_pagina_real = 0

    for bloco in paginas:
        texto = bloco.strip()
        if not texto:
            continue

        if "LIVRO REGISTRO DE APURAÇÃO DO ICMS" not in texto.upper():
            continue

        num_pagina_real += 1

        if TITULO_RESUMO in texto.upper():
            continue

        meta = extrair_cabecalho_pagina(texto)

        if not filial:
            filial = encontrar_filial_por_txt(
                caminho_txt,
                meta["CNPJ"],
                meta["IE"]
            )

        divisao = resolver_divisao_por_filial_e_pasta(caminho_txt, filial, mapa_divisao)

        linhas = texto.splitlines()
        tipo_atual = None

        for linha in linhas:
            linha_limpa = linha.strip()

            if not linha_limpa:
                continue

            novo_tipo = linha_indica_secao_cfop(linha_limpa)
            if novo_tipo:
                tipo_atual = novo_tipo
                continue

            if deve_ignorar_linha_cfop(linha_limpa):
                continue

            if eh_linha_numerica_tabela(linha_limpa):
                dados_linha = parse_linha_tabela(linha_limpa)
                if dados_linha:
                    dados_linha["Imposto Creditado"] = aplicar_sinal_entrada(
                        dados_linha["Imposto Creditado"], tipo_atual
                    )

                    registros.append({
                        "Arquivo PDF": arquivo_txt,
                        "Filial": filial,
                        "Divisão": divisao,
                        "Página": num_pagina_real,
                        "Firma": meta["Firma"],
                        "IE": meta["IE"],
                        "CNPJ": meta["CNPJ"],
                        "Período": meta["Período"],
                        "Tipo": tipo_atual,
                        "Status": "Entrada / Saída",
                        "Mapeamento": mapa_cfop.get(str(dados_linha["CFOP"]).strip(), ""),
                        **dados_linha
                    })

    return registros

def processar_txt_resumo(caminho_txt, mapa_divisao):
    registros = []
    arquivo_txt = os.path.basename(caminho_txt)
    filial = extrair_filial_do_arquivo(caminho_txt)

    try:
        texto_total = ler_txt_com_fallback(caminho_txt)
    except Exception:
        return registros

    paginas = re.split(
        r"(?=LIVRO REGISTRO DE APURAÇÃO DO ICMS - RAICMS - MODELO P9|LIVRO REGISTRO DE APURAÇÃO DO ICMS - RAICMS - MODELO P)",
        texto_total
    )

    seq = 0
    num_pagina_real = 0

    for bloco in paginas:
        texto = bloco.strip()
        if not texto:
            continue

        if "LIVRO REGISTRO DE APURAÇÃO DO ICMS" not in texto.upper():
            continue

        num_pagina_real += 1

        if TITULO_RESUMO not in texto.upper():
            continue

        meta = extrair_cabecalho_pagina(texto)

        if not filial:
            filial = encontrar_filial_por_txt(
                caminho_txt,
                meta["CNPJ"],
                meta["IE"]
            )

        divisao = resolver_divisao_por_filial_e_pasta(caminho_txt, filial, mapa_divisao)

        linhas = texto.splitlines()
        secao_atual = None

        for linha in linhas:
            linha_limpa = linha.strip()

            if not linha_limpa:
                continue

            nova_secao = linha_indica_secao_resumo(linha_limpa)
            if nova_secao:
                secao_atual = nova_secao
                continue

            if deve_ignorar_linha_resumo(linha_limpa):
                continue

            dados = parse_linha_resumo(linha_limpa)
            if dados:
                seq += 1
                tipo = definir_tipo_resumo(secao_atual, dados["Descrição"])
                dados["Somas"] = aplicar_sinal_entrada(dados["Somas"], tipo)
                status = definir_status_por_descricao(dados["Descrição"])

                registros.append({
                    "_seq": seq,
                    "Arquivo PDF": arquivo_txt,
                    "Filial": filial,
                    "Divisão": divisao,
                    "Página": num_pagina_real,
                    "Firma": meta["Firma"],
                    "IE": meta["IE"],
                    "CNPJ": meta["CNPJ"],
                    "Período": meta["Período"],
                    "Seção": secao_atual,
                    "Tipo": tipo,
                    "Status": status,
                    **dados
                })

    return registros

# =========================================================
# CONFERÊNCIA
# =========================================================
def montar_conferencia(df_cfop, df_resumo):
    partes = []

    if not df_cfop.empty:
        tmp_cfop = df_cfop[["Divisão", "Filial", "Status", "Imposto Creditado"]].copy()
        tmp_cfop = tmp_cfop.rename(columns={"Imposto Creditado": "Montante"})
        partes.append(tmp_cfop)

    if not df_resumo.empty:
        descricoes_desconsiderar = {
            "ENTRADAS",
            "SAIDAS",
            "SUB TOTAL",
            "SALDO DEVEDOR",
            "TOTAL"
        }

        tmp_resumo = df_resumo.copy()
        tmp_resumo["_DESC_NORM"] = tmp_resumo["Descrição"].apply(normalizar_texto)
        tmp_resumo = tmp_resumo[~tmp_resumo["_DESC_NORM"].isin(descricoes_desconsiderar)]

        tmp_resumo = tmp_resumo[["Divisão", "Filial", "Status", "Somas"]].copy()
        tmp_resumo = tmp_resumo.rename(columns={"Somas": "Montante"})
        partes.append(tmp_resumo)

    if not partes:
        return pd.DataFrame(columns=[
            "Divisão", "Filial", "Entrada / Saída",
            "ICMS a Recolher ou Recuperar", "Soma Total"
        ])

    base = pd.concat(partes, ignore_index=True)
    base["Montante"] = pd.to_numeric(base["Montante"], errors="coerce").fillna(0).round(2)

    conf = (
        base.groupby(["Divisão", "Filial", "Status"], dropna=False, as_index=False)["Montante"]
        .sum()
        .pivot(index=["Divisão", "Filial"], columns="Status", values="Montante")
        .fillna(0)
        .reset_index()
    )

    conf.columns.name = None

    for col in ["Entrada / Saída", "ICMS a Recolher ou Recuperar"]:
        if col not in conf.columns:
            conf[col] = 0.0

    conf["Entrada / Saída"] = pd.to_numeric(
        conf["Entrada / Saída"], errors="coerce"
    ).fillna(0).round(2)

    conf["ICMS a Recolher ou Recuperar"] = pd.to_numeric(
        conf["ICMS a Recolher ou Recuperar"], errors="coerce"
    ).fillna(0).round(2)

    # Total Geral = Apuração Vivo + Livro Entrada / Saída
    conf["Soma Total"] = (
        conf["Entrada / Saída"] + conf["ICMS a Recolher ou Recuperar"]
    ).round(2)

    # limpa resíduos tipo -0.00 / 0.00 quebrado
    for col in ["Entrada / Saída", "ICMS a Recolher ou Recuperar", "Soma Total"]:
        conf.loc[conf[col].abs() < 0.005, col] = 0.0

    conf = conf[[
        "Divisão",
        "Filial",
        "Entrada / Saída",
        "ICMS a Recolher ou Recuperar",
        "Soma Total"
    ]]
    return conf

# =========================================================
# EXCEL
# =========================================================
def aplicar_formatacao_excel(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)

    try:
        moeda_style = NamedStyle(name="moeda_style_raicms", number_format='R$ #,##0.00')
        wb.add_named_style(moeda_style)
    except Exception:
        pass

    abas_colunas_monetarias = {
        "Consolidado": [
            "Valores Contábeis",
            "Base de Cálculo",
            "Imposto Creditado",
            "Isentas ou não Trib.",
            "Outras"
        ],
        "Resumo_Apuracao": [
            "Coluna Auxiliar",
            "Somas"
        ],
        "Conferência": [
            "Entrada / Saída",
            "ICMS a Recolher ou Recuperar",
            "Soma Total"
        ]
    }

    for nome_aba, colunas_monetarias in abas_colunas_monetarias.items():
        if nome_aba not in wb.sheetnames:
            continue

        ws = wb[nome_aba]
        headers = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)}

        for nome_col in colunas_monetarias:
            col_idx = headers.get(nome_col)
            if not col_idx:
                continue

            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = 'R$ #,##0.00'

    wb.save(caminho_arquivo)

def salvar_cache_raicms(arquivo_final, arquivos_gerados=None):
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    meta = {
        "arquivo_final": str(arquivo_final),
        "arquivos_gerados": [str(x) for x in (arquivos_gerados or [])],
    }
    with open(CACHE_RAICMS_META, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

# =========================================================
# MOTOR DE PROCESSAMENTO
# =========================================================
def processar_raicms(pasta_pdfs, pasta_destino, progress_callback=None):
    t0 = time.time()

    mapa_divisao = carregar_mapa_divisao()
    mapa_cfop = carregar_mapa_cfop()

    dados_cfop = []
    dados_resumo = []
    nao_reconhecidos = []
    arquivos_sem_dados = []

    arquivos_pdf = []
    arquivos_txt = []

    for raiz, _, files in os.walk(pasta_pdfs):
        for f in files:
            caminho_completo = os.path.join(raiz, f)

            if f.lower().endswith(".pdf"):
                arquivos_pdf.append(caminho_completo)
            elif f.lower().endswith(".txt"):
                arquivos_txt.append(caminho_completo)

    arquivos_pdf = sorted(arquivos_pdf)
    arquivos_txt = sorted(arquivos_txt)

    if not arquivos_pdf and not arquivos_txt:
        raise FileNotFoundError("Nenhum PDF ou TXT encontrado na pasta selecionada.")

    pastas_com_txt = {os.path.dirname(x) for x in arquivos_txt}

    arquivos_fonte = []

    # TXTs têm prioridade total
    arquivos_fonte.extend(arquivos_txt)

    # PDFs só entram se a pasta não tiver TXT
    for caminho_pdf in arquivos_pdf:
        pasta_pdf = os.path.dirname(caminho_pdf)
        if pasta_pdf not in pastas_com_txt:
            arquivos_fonte.append(caminho_pdf)

    arquivos_fonte = sorted(arquivos_fonte)
    total = len(arquivos_fonte)

    for i, caminho_fonte in enumerate(arquivos_fonte, start=1):
        if progress_callback:
            progress_callback("processando_pdf", i, total, os.path.basename(caminho_fonte))

        eh_txt = caminho_fonte.lower().endswith(".txt")
        filial_tmp = extrair_filial_do_arquivo(caminho_fonte)

        if not filial_tmp:
            if eh_txt:
                try:
                    texto_tmp = ler_txt_com_fallback(caminho_fonte)
                    meta_tmp = extrair_cabecalho_pagina(texto_tmp)

                    filial_tmp = encontrar_filial_priorizando_txts_da_pasta(
                        caminho_fonte,
                        meta_tmp["CNPJ"],
                        meta_tmp["IE"]
                    )
                except Exception:
                    pass
            else:
                try:
                    with pdfplumber.open(caminho_fonte) as pdf_tmp:
                        for pagina_tmp in pdf_tmp.pages:
                            texto_tmp = pagina_tmp.extract_text() or ""

                            if TITULO_ALVO not in texto_tmp:
                                continue

                            meta_tmp = extrair_cabecalho_pagina(texto_tmp)

                            filial_tmp = encontrar_filial_priorizando_txts_da_pasta(
                                caminho_fonte,
                                meta_tmp["CNPJ"],
                                meta_tmp["IE"]
                            )

                            if filial_tmp:
                                break
                except Exception:
                    pass

        divisao_tmp = resolver_divisao_por_filial_e_pasta(caminho_fonte, filial_tmp, mapa_divisao)

        if not filial_tmp or not divisao_tmp:
            pasta_fonte = os.path.dirname(caminho_fonte)

            try:
                qtd_arquivos_mesma_pasta = len([
                    f for f in os.listdir(pasta_fonte)
                    if f.lower().endswith(".pdf") or f.lower().endswith(".txt")
                ])
            except Exception:
                qtd_arquivos_mesma_pasta = 0

            if qtd_arquivos_mesma_pasta > 1:
                nao_reconhecidos.append({
                    "Arquivo PDF": os.path.basename(caminho_fonte),
                    "Caminho Pasta": pasta_fonte,
                    "Filial Reconhecida": filial_tmp,
                    "Divisão Reconhecida": divisao_tmp,
                    "Qtd PDFs na Pasta": qtd_arquivos_mesma_pasta
                })

        qtd_cfop_antes = len(dados_cfop)
        qtd_resumo_antes = len(dados_resumo)

        if eh_txt:
            dados_cfop.extend(processar_txt_cfop(caminho_fonte, mapa_divisao, mapa_cfop))
            dados_resumo.extend(processar_txt_resumo(caminho_fonte, mapa_divisao))
        else:
            dados_cfop.extend(processar_pdf_cfop(caminho_fonte, mapa_divisao, mapa_cfop))
            dados_resumo.extend(processar_pdf_resumo(caminho_fonte, mapa_divisao))

        qtd_cfop_depois = len(dados_cfop)
        qtd_resumo_depois = len(dados_resumo)

        if qtd_cfop_depois == qtd_cfop_antes and qtd_resumo_depois == qtd_resumo_antes:
            arquivos_sem_dados.append({
                "Arquivo Fonte": os.path.basename(caminho_fonte),
                "Caminho Completo": caminho_fonte,
                "Tipo Arquivo": "TXT" if eh_txt else "PDF",
                "Filial Reconhecida": filial_tmp,
                "Divisão Reconhecida": divisao_tmp
            })

    df_cfop = pd.DataFrame(dados_cfop) if dados_cfop else pd.DataFrame(columns=[
        "Arquivo PDF", "Filial", "Divisão", "Página", "Firma", "IE", "CNPJ", "Período", "Tipo", "Status",
        "CFOP", "Mapeamento", "Valores Contábeis", "Base de Cálculo", "Imposto Creditado",
        "Isentas ou não Trib.", "Outras"
    ])
    df_resumo = pd.DataFrame(dados_resumo) if dados_resumo else pd.DataFrame(columns=[
        "Arquivo PDF", "Filial", "Divisão", "Página", "Firma", "IE", "CNPJ", "Período",
        "Seção", "Tipo", "Status", "Número", "Descrição", "Coluna Auxiliar", "Somas"
    ])

    if not df_cfop.empty:
        cols_ordem_cfop = [
            "Arquivo PDF", "Filial", "Divisão", "Página", "Firma", "IE", "CNPJ", "Período", "Tipo", "Status",
            "CFOP", "Mapeamento", "Valores Contábeis", "Base de Cálculo", "Imposto Creditado",
            "Isentas ou não Trib.", "Outras"
        ]
        df_cfop = df_cfop[cols_ordem_cfop]

    if not df_resumo.empty:
        df_resumo = df_resumo.sort_values(by=["Arquivo PDF", "_seq"], kind="stable")
        cols_ordem_resumo = [
            "Arquivo PDF", "Filial", "Divisão", "Página", "Firma", "IE", "CNPJ",
            "Período", "Seção", "Tipo", "Status", "Número", "Descrição", "Coluna Auxiliar", "Somas"
        ]
        df_resumo = df_resumo[cols_ordem_resumo]

    df_conferencia = montar_conferencia(df_cfop, df_resumo)

    if df_cfop.empty and df_resumo.empty:
        raise ValueError("Nenhum dado útil foi encontrado nos PDFs.")

    pares = set()

    if not df_cfop.empty:
        pares.update(tuple(x) for x in df_cfop[["Período"]].drop_duplicates().values.tolist())

    if not df_resumo.empty:
        pares.update(tuple(x) for x in df_resumo[["Período"]].drop_duplicates().values.tolist())

    arquivos_gerados = []

    total_periodos = max(1, len(pares))
    for idx, (periodo,) in enumerate(sorted(pares, key=lambda x: str(x[0])), start=1):
        if progress_callback:
            progress_callback("gerando_excel", idx, total_periodos, f"Período {periodo}")

        periodo_txt = formatar_periodo(periodo if pd.notna(periodo) else "SEM_PERIODO")
        nome_arquivo = f"RAICMS - {periodo_txt}.xlsx"
        caminho_saida = os.path.join(pasta_destino, nome_arquivo)

        df_cfop_grupo = df_cfop[(df_cfop["Período"] == periodo)] if not df_cfop.empty else pd.DataFrame()
        df_resumo_grupo = df_resumo[(df_resumo["Período"] == periodo)] if not df_resumo.empty else pd.DataFrame()
        df_conferencia_grupo = montar_conferencia(df_cfop_grupo, df_resumo_grupo)

        with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
            if not df_cfop_grupo.empty:
                df_cfop_grupo.to_excel(writer, index=False, sheet_name="Consolidado")
            else:
                pd.DataFrame(columns=[
                    "Arquivo PDF", "Filial", "Divisão", "Página", "Firma", "IE", "CNPJ", "Período", "Tipo", "Status",
                    "CFOP", "Mapeamento", "Valores Contábeis", "Base de Cálculo", "Imposto Creditado",
                    "Isentas ou não Trib.", "Outras"
                ]).to_excel(writer, index=False, sheet_name="Consolidado")

            if not df_resumo_grupo.empty:
                df_resumo_grupo.to_excel(writer, index=False, sheet_name="Resumo_Apuracao")
            else:
                pd.DataFrame(columns=[
                    "Arquivo PDF", "Filial", "Divisão", "Página", "Firma", "IE", "CNPJ",
                    "Período", "Seção", "Tipo", "Status", "Número", "Descrição", "Coluna Auxiliar", "Somas"
                ]).to_excel(writer, index=False, sheet_name="Resumo_Apuracao")

            df_conferencia_grupo.to_excel(writer, index=False, sheet_name="Conferência")

        aplicar_formatacao_excel(caminho_saida)
        arquivos_gerados.append(caminho_saida)

    if progress_callback:
        progress_callback("gerando_excel", total_periodos, total_periodos, "Consolidado geral")

    caminho_consolidado = os.path.join(pasta_destino, "RAICMS - Consolidado Geral.xlsx")
    with pd.ExcelWriter(caminho_consolidado, engine="openpyxl") as writer:
        df_cfop.to_excel(writer, index=False, sheet_name="Consolidado")
        df_resumo.to_excel(writer, index=False, sheet_name="Resumo_Apuracao")
        df_conferencia.to_excel(writer, index=False, sheet_name="Conferência")

    aplicar_formatacao_excel(caminho_consolidado)
    arquivos_gerados.append(caminho_consolidado)

    caminho_nao_reconhecidos = None

    if nao_reconhecidos:
        caminho_nao_reconhecidos = os.path.join(
            pasta_destino,
            "RAICMS - PDFs sem filial reconhecida.xlsx"
        )

        df_nao_reconhecidos = pd.DataFrame(nao_reconhecidos).drop_duplicates()
        df_nao_reconhecidos.to_excel(caminho_nao_reconhecidos, index=False)
        arquivos_gerados.append(caminho_nao_reconhecidos)

    caminho_sem_dados = None

    if arquivos_sem_dados:
        caminho_sem_dados = os.path.join(
            pasta_destino,
            "RAICMS - Arquivos sem dados extraídos.xlsx"
        )

        df_sem_dados = pd.DataFrame(arquivos_sem_dados).drop_duplicates()
        df_sem_dados.to_excel(caminho_sem_dados, index=False)
        arquivos_gerados.append(caminho_sem_dados)

    salvar_cache_raicms(
        arquivo_final=caminho_consolidado,
        arquivos_gerados=arquivos_gerados,
    )
    return {
        "arquivos_pdf": len(arquivos_pdf),
        "linhas_cfop": len(df_cfop),
        "linhas_resumo": len(df_resumo),
        "tempo_total": round(time.time() - t0, 2),
        "arquivo_final": caminho_consolidado,
        "arquivos_gerados": arquivos_gerados,
        "arquivo_nao_reconhecidos": caminho_nao_reconhecidos,
        "arquivo_sem_dados": caminho_sem_dados
    }